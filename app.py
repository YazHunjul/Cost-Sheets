import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import shutil
import os
from docx import Document
from docx.shared import Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docxtpl import DocxTemplate
import math
import zipfile
import re

# Add this dictionary at the top of the file, after imports
contacts = {
    'Marc Byford': '(07974 403322)',
    'Karl Nicholson': '(07791 397866)',
    'Dan Butler': '(07703 729686)',
    'Chris Mannus': '(07870 263280)',
    'Dean Griffiths': '(07814 784352)',
    'Kent Phillips': '(07949 016501)'
}

estimators = {
    'Simon Still': 'Lead Estimator',
    'Nick Soton': 'Estimator',
    'Chris Davis': 'Estimator'
}

# Company addresses dictionary
company_addresses = {
    "Airedale Group (Bradford)": "Victoria Road\nPedeshill\nBradford BD23 2BN",
    "Airedale Group (Lutterworth)": "1 St Johns Business Park\nRugby Road\nLutterworth LE17 4HB",
    "Court Catering Equipment Ltd": "Unit 1, Acton Vale Ind. Park,\nCowley Road\nLondon W3 7XA",
    "Humble Arnold Associates": "Farriers House\nFarriers Close\nCodicote\nHertfordshire\nSG4 8DU",
    "Berkeley Projects UK Ltd": "17 Ewell Road\nCheam\nSurrey\nSM3 8DD",
    "Chapman Ventilation Ltd": "15-20 Woodfield Rd\nWelwyn Garden City\nHerts AL7 1JQ",
    "SG Group": "Aspen Way\nPaignton\nDevon TQ4 7QR",
    "ABM Catering for Leisure Ltd": "Algate House\nClydesmuir Rd Ind Est\nCardiff CF24 2QS",
    "C&C Catering Equipment Ltd": "1 Smithy Farm\nChapel Lane\nSaighton\nChester CH3 6EW",
    "Hallmark Kitchens Ltd": "South Barn\nCrockham Farm\nEdenbridge\nKent\nTN8 6SR",
    "Cabiola Foodservice Equipment": "The Bake House\nNarborough Wood Park\nDesford Road\nEnderby\nLeicester LE19 4XT",
    "Ceba Catering Services": "Unit 27, Eastville Close\nEastern Avenue Trading Estate\nGloucester\nGL4 3SJ",
    "Design Installation Service Ltd": "4 Gainsborough House\n42 / 44 Bath Road\nCheltenham\nGL53 7HJ",
    "Nelson Bespoke Commercial Kitchens": "Unit 1\nRowley Industrial Park\nRoslin Road\nActon\nLondon W3 8BH",
    "Airflow Cooling Ltd": "132 Rutland Road\nSheffield S3 9PP",
    "Spectrum Contracts Ltd.": "Unit 11 Dorcan Business Village\nMurdoch Road\nDorcan\nSwindon\nSN3 5HY",
    "CCE Group Ltd": "Unit 1 Bentley Farm\nOld Church Hill\nChingon Hills\nBasildon SS16 6HZ",
    "Shine Food Machinery Ltd": "New Quay Road\nStephenson Street Ind. Est.\nNewport NP19 4FL",
    "HCE Foodservice Equipment Ltd": "School Lane\nChandlers Ford\nEastleigh\nHants SO53 4DG",
    "Vision Commercial Kitchens Limited": "Unit A1, Axis Point,\nHill Top Road,\nHeywood\nLancs OL10 2RQ",
    "Fast Food Systems Ltd.": "Unit 1\nHeadley Park 9\nHeadley Road East\nWoodley\nReading\nBerkshire RG5 4SQ",
    "C. Caswell (Eng) Services Ltd": "Knowsley Road Ind. Est.\nHaslingden\nRossendale\nLancs BB4 4RX",
    "VSS Ltd": "Building 2 ThermoAir Site\nAthy Road\nCarlow\nIreland R93 K635",
    "Reco-Air": "Newmarket 24, Centrix,\nKeys Business Village\nCannock\nStaffordshire\nWS12 2HA",
    "Gratte Brothers Ltd.": "3 Crompton Road\nStevenage\nHertfordshire\nSG1 2XP",
    "HCE Foodservice Equipment Ltd": "School Lane\nChandlers Ford Id Est.\nEastleigh\nHants SO53 4DG",
    "Salix Stainless Steel Production House": "Chester Hall Lane\nBasildon\nEssex SS14 3BG",
    "Stangard Design Solutions Ltd.": "The Dairy\nWolvey Lodge Business Centre\nWolvey\nWarwickshire\nLE10 3HB",
    "AFR Air Conditioning Ltd": "14 Kingsdown Road\nSwindon\nWiltshire\nSN25 6PB",
    "Sigma Catering Equipment": "Unit 4\nPotter Street\nWallsend\nNE28 6LS",
    "CHR Equipment Ltd.": "Astar House\nFourries Bridge\nPreston\nPR5 6GS",
    "MITIE Engineering Services (Retail) Ltd.": "The Millennium Centre\nM4 Crosby Way\nFarnham\nSurrey GU9 7XX",
    "Catering Design Services": "4 Waterside Commerce Park\nTraffod Park\nManchester\nM17 1WD",
    "Salix Stainless Steel Applications": "Production House\nChester Hall Lane\nBasildon\nEssex SS14 3BG",
    "Galgorm Group": "Galgorm Industrial Estate\n7 Corbally Road\nBallymena BT42 1JQ",
    "Michael J Lonsdale": "Unit 1Langley Quay\nWaterside Drive\nLangley\nSlough\nBerkshire\nSL3 6EY",
    "Advance Catering Equipment": "Dantem House\nBlackburn Road\nHoughton Regis\nDunstable\nLU5 5BQ",
    "MITIE Engineering Services (South West) Ltd": "5 Hanover Court\nMatford Close\nMatford Business Park\nExeter\nEX2 8QJ",
    "A C V R Services Ltd": "Unit 4\nNewtown Grange Farm Bu siness Park Orchard\nLeicester\nLE8 8FL",
    "Lockhart Catering Equipment": "6 The Astra Centre\nEdinburgh Way\nHarlow\nEssex\nCM20 2BN",
    "Insitu Fabrications": "Unit19\nButterfield Ind. Est.\nNewtongrange\nBonnyrigg EH19 3JQ",
    "Stephens Catering Equipment": "205 Carnabanaugh Road\nDoughandshane\nBallymena\nCo. Antrim BT42 4NY",
    "Into Design Ltd": "16 Richards Rise\nOxted\nSurrey\nRH8 0TS",
    "Merrick Contracts Ltd": "Bird Point Mill\nKing Henry's Drive\nNew Addington\nSurrey\nCR0 0AE",
    "CKE Service Ltd": "13 High Street\nThames Ditton\nSurrey\nKT7 0SN",
    "Dentons Catering Equipment": "2/4 Clapham High Street\nLondon\nSW4 7UT",
    "Ferro Design Ltd": "Blay's House\nChurchfield Road\nChalfont St Peter\nBucks\nSL9 9EW",
    "Four Seasons Air Conditioning Suppliers Ltd": "Stadium Works\nSedgley Street\nWolverhampton\nWest Midlands\nWV2 3AJ",
    "Stephens Catering Equipment": "Block F, Unit 9\nMaynooth Business Campus\nMaynooth, Co. Kildare.",
    "Reco-Air": "14 Heritage Park\nHayes Way\nCannock\nStaffordshire\nWS11 7LT",
    "Pan-Euro Environmental": "Unit 2, Group House\nAlbon Way\nReigate\nRH2 7JY",
    "Caswell Engineering": "Knowsley Road Ind. Est.\nHaslingden\nRossendale\nLancs BB4 4RR",
    "C A Sothers Ltd": "156 Hockey Hill\nBirmingham\nB18 5AN",
    "RDA": "5 Apollo Court\nMonkton Business Park South\nTyne & Wear NE31 2ES",
    "KCCJ Ltd.": "The Old Granary\nCourt Lodge Farm\nLambden Hill\nDA2 7QY",
    "Contract Catering Equipment": "Unit1, Bentley Farm,\nOld Church Hill, Chingon Hills,\nEssex SS16 6HZ",
    "Fellerman Partnership Ltd": "74 Kimberley Road\nPortsmouth PO4 9NS",
    "Tricon Foodservice Consultants": "St James House\n27-43 Eastern Road\nRomford",
    "Technical Services Ref & A/C Ltd": "Arlingston House\n32 Boundary Road\nNewbury",
    "Sefton Horn Winch": "The Stables\nHome Farm Business Units\nRiverside\nEynsford",
    "Carford Group": "1-4 Mitchell Road\nFernside Park\nFerndown Industrial Estate\nFerndown\nDorset BH21 7SG",
    "Gratte Bros. Catering Equipment Ltd": "3 Crompton Road\nStevenage\nHerts SG1 2XP",
    "Scobie & McIntosh": "15 Brewster Square\nBrucefield Industry Park\nLivingston\nWest Lothian\nEH54 9BJ",
    "SVS Ltd": "Unit 3\nGreencroft Ind. Est.\nAnnfield Plain\nStanley\nCo Durham\nDH9 7YB",
    "Kitchequip": "Canal View\nWaterside Business Park\nNew Lane\nBurscough L40 8JX",
    "Space Catering": "Barnwood Point\nCorinium Avenue\nGlos GL4 3HX",
    "GS Catering Equipment Ltd": "Aspen Way\nPaignton\nDevon TQ4 7QR",
    "GastroNorth Ltd": "Merlin House\nTeam Valley Trading Est\nPrinces Park\nGateshead\nNE11 0NF",
    "Promart Manufacturing Ltd": "2B Caddick Road\nKnowsley Business Park\nPrescot\nMerseyside L34 9HP",
    "Bob Sackett Commercial Catering Projects Ltd": "9 Chesterford Court\nLondon Road\nGreat Chesterford\nEssex CB10 1PF",
    "Vision Commercial Kitchens Limited": "Unit A1, Axis Point,\nHill Top Road,\nHeywood\nLancs OL10 2RQ",
    "Court Catering Equipment Ltd": "8 Cowley Road\nLondon W3 7XA",
    "Main Contract Services Ltd": "Alexandra House\nStation Road\nGrangemouth FK3 8DG",
    "CDS-Wilman": "4 Waterside Commerce Park\nTrafford Park\nManchester M17 1W",
    "AFR Air Conditioning": "14 Kingsdown Road\nSwindon\nWilts SN25 6PB",
    "Western Blueprint Ltd": "Unit B2, 1st Floor\nTrym House Business Centre\nForest Road\nKingswood\nBristol BS15 8DH",
    "Edge Design": "Unit 3 Wiston House\nWiston Avenue\nWorthing\nWest Sussex BN14 7QL",
    "Chapman Ventilation": "15 - 20 Woodfield Road\nWelwyn Garden City\nHertfordshire\nAL7 1JQ",
    "Michael J Lonsdale": "Unit 1 Langley Quay\nWaterside Drive\nBerkshire SL3 6EY",
    "RecoAir": "14 Heritage Park\nHayes Way\nCannock\nStaffordshire WS11 7LT",
    "Humble Arnold": "Farriers House,\nFarriers Close,\nCodicote\nHerts SG4 8DU",
}

# Add this at the top of the file with other constants
VALID_CANOPY_MODELS = [
    'KVF', 'UVF', 'CMWF', 'CMWI', 'KVI', 'UVI', 'CXW', 'KVS', 'UVS', 'KSW'
]  # Add any other valid canopy models

# Add this lookup table after other constants at the top of the file
RECOAIR_SPECS = {
    # Vertical models
    'RAH0.5V': {'p_drop': 1050, 'motor': '2.2', 'weight': 436},
    'RAH0.8V': {'p_drop': 1050, 'motor': '2.2', 'weight': 470},
    'RAH1.0V': {'p_drop': 1050, 'motor': '4.7', 'weight': 572},
    'RAH1.5V': {'p_drop': 1050, 'motor': '4.7', 'weight': 820},
    'RAH2.0V': {'p_drop': 1050, 'motor': '5.25', 'weight': 974},
    'RAH2.5V': {'p_drop': 1050, 'motor': '5.25', 'weight': 1170},
    'RAH3.0V': {'p_drop': 1050, 'motor': '5.25', 'weight': 1210},
    
    # Horizontal models
    'RAH0.5': {'p_drop': 1050, 'motor': '2.2', 'weight': 385},
    'RAH0.8': {'p_drop': 1050, 'motor': '2.2', 'weight': 415},
    'RAH1.0': {'p_drop': 1050, 'motor': '4.7', 'weight': 542},
    'RAH1.5': {'p_drop': 1050, 'motor': '4.7', 'weight': 765},
    'RAH2.0': {'p_drop': 1050, 'motor': '5.25', 'weight': 884},
    'RAH2.5': {'p_drop': 1050, 'motor': '5.25', 'weight': 1093},
    'RAH3.0': {'p_drop': 1050, 'motor': '5.25', 'weight': 1210},
    'RAH3.5': {'p_drop': 1050, 'motor': '5.25', 'weight': 1395},
    'RAH4.0': {'p_drop': 1050, 'motor': '5.25', 'weight': 1500}
}

# Add this function after the imports
def create_project_sidebar(project_data):
    """Create a sidebar showing project progress and structure"""
    with st.sidebar:
        st.header("📋 Project Overview")
        
        # Show General Info Status
        st.subheader("General Information")
        
        # Check required fields
        required_fields = {
            'Project Name': project_data.get('Project Name', ''),
            'Project Number': project_data.get('Project Number', ''),
            'Customer': project_data.get('Customer', ''),
            'Company': project_data.get('Company', ''),
            'Location': project_data.get('Location', ''),
            'Address': project_data.get('Address', ''),
            'Sales Contact': project_data.get('Sales Contact', ''),
            'Estimator': project_data.get('Estimator', '')
        }
        
        # Show status of each required field
        for field, value in required_fields.items():
            if value and value != "Select...":
                st.markdown(f"✅ {field}")
            else:
                st.markdown(f"❌ {field} *required*")
        
        # Show Levels and Areas
        st.markdown("---")
        st.subheader("Project Structure")
        
        levels = project_data.get('Levels', [])
        if not levels:
            st.info("No levels added yet")
            return
            
        for level in levels:
            if isinstance(level, dict):  # Make sure level is a dictionary
                level_name = level.get('level_name', '')
                areas = level.get('areas', [])
                
                if level_name:
                    st.markdown(f"### 🏢 {level_name}")
                    
                    if not areas:
                        st.markdown("- *No areas added*")
                        continue
                        
                    for area in areas:
                        if isinstance(area, dict):  # Make sure area is a dictionary
                            area_name = area.get('area_name', '')
                            canopies = area.get('canopies', [])
                            
                            if area_name:
                                st.markdown(f"#### 📍 {area_name}")
                                
                                if not canopies:
                                    st.markdown("- *No canopies added*")
                                    continue
                                    
                                valid_canopies = [
                                    canopy for canopy in canopies 
                                    if isinstance(canopy, dict) and  # Make sure canopy is a dictionary
                                    canopy.get('reference_number') and 
                                    canopy.get('model') != "Select..." and 
                                    canopy.get('configuration') != "Select..."
                                ]
                                
                                if valid_canopies:
                                    for canopy in valid_canopies:
                                        ref = canopy.get('reference_number', '')
                                        model = canopy.get('model', '')
                                        config = canopy.get('configuration', '')
                                        st.markdown(f"- 🔹 {ref}: {model} ({config})")
                                        
                                        # Show wall cladding if present
                                        if (isinstance(canopy.get('wall_cladding'), dict) and 
                                            canopy['wall_cladding'].get('type') and 
                                            canopy['wall_cladding']['type'] != "Select..."):
                                            positions = '/'.join(canopy['wall_cladding'].get('positions', []))
                                            st.markdown(f"  - 🧱 Wall Cladding: {positions}")
                                else:
                                    st.markdown("- *No valid canopies configured*")

def create_general_info_form():
    # Initialize session state if needed
    if 'project_data' not in st.session_state:
        st.session_state.project_data = {
            'Project Name': '',
            'Project Number': '',
            'Customer': '',
            'Company': '',
            'Location': '',
            'Address': '',
            'Sales Contact': '',
            'Estimator': '',
            'Levels': []
        }
    
    st.title("📋 Project Information")
    
    # General Information Section
    col1, col2, col3 = st.columns(3)
    
    with col1:
        project_name = st.text_input("Project Name")
        st.session_state.project_data['Project Name'] = project_name
    with col2:
        project_number = st.text_input("Project Number")
        st.session_state.project_data['Project Number'] = project_number
    with col3:
        date = st.date_input("Date", datetime.now())
        st.session_state.project_data['Date'] = date.strftime("%d/%m/%Y")
    
    col4, col5, col6 = st.columns(3)
    
    with col4:
        customer = st.text_input("Customer")
        st.session_state.project_data['Customer'] = customer
    with col5:
        company = st.text_input("Company")
        st.session_state.project_data['Company'] = company
    with col6:
        location = st.text_input("Location")
        st.session_state.project_data['Location'] = location
    
    col7, col8, col9 = st.columns(3)
    
    with col7:
        # Address dropdown with all company addresses and custom option
        address_options = ["Select..."] + list(company_addresses.keys()) + ["Custom Address"]
        selected_address = st.selectbox("Company Address", address_options)
        
        # Handle the address selection
        if selected_address == "Custom Address":
            custom_address = st.text_area("Enter Custom Address", height=100)
            address = custom_address
        elif selected_address != "Select...":
            # Display the selected company address
            address = company_addresses[selected_address]
            st.text_area("Address", address, height=100, disabled=True)
        else:
            address = ""
        
        st.session_state.project_data['Address'] = address
    with col8:
        sales_contact = st.selectbox("Sales Contact", ["Select..."] + list(contacts.keys()))
        st.session_state.project_data['Sales Contact'] = sales_contact
    with col9:
        estimator = st.selectbox("Estimator", ["Select..."] + list(estimators.keys()))
        st.session_state.project_data['Estimator'] = estimator
    
    cost_sheet = st.selectbox("Select Cost Sheet to Report", ["Select...", "Canopy", "Reco Air"])
    st.session_state.project_data['Cost Sheet'] = cost_sheet
    
    # Levels Configuration
    if cost_sheet != "Select...":
        st.markdown("---")
        st.subheader("Levels Configuration")
        
        num_levels = st.number_input("Enter Number of Levels", min_value=1, value=1, step=1)
        
        levels_data = []
        
        for level_idx in range(num_levels):
            with st.expander(f"Level {level_idx + 1}", expanded=False):
                level_name = st.text_input(f"Enter Level {level_idx + 1} Name", 
                                         key=f"level_name_{level_idx}")
                
                if level_name:
                    num_areas = st.number_input(
                        f"Enter the number of areas in {level_name}",
                        min_value=1,
                        value=1,
                        step=1,
                        key=f"num_areas_{level_idx}"
                    )
                    
                    areas_data = []
                    
                    for area_idx in range(num_areas):
                            area_name = st.text_input(
                                f"Enter area {area_idx + 1} Name",
                                key=f"area_name_{level_idx}_{area_idx}"
                            )
                            
                            if area_name:
                                # Show different options based on cost sheet type
                                if cost_sheet == "Canopy":
                                    # Add UV-C Control Schedule radio button for the area
                                    include_uvc = st.radio(
                                        "Include UV-C Control Schedule",
                                        options=["No", "Yes"],
                                        key=f"uvc_{level_idx}_{area_idx}"
                                    )
                                    
                                    # Add RECOAIR system radio button for the area
                                    include_recoair = st.radio(
                                        "Include RECOAIR System",
                                        options=["No", "Yes"],
                                        key=f"recoair_{level_idx}_{area_idx}"
                                    )
                                    
                                    # Add SDU option radio button for the area
                                    include_sdu = st.radio(
                                        "Include SDU",
                                        options=["No", "Yes"],
                                        key=f"sdu_{level_idx}_{area_idx}"
                                    )
                                    
                                    num_canopies = st.number_input(
                                        f"Enter Number of Canopies in {area_name}",
                                        min_value=0,
                                        value=0,
                                        step=1,
                                        key=f"num_canopies_{level_idx}_{area_idx}"
                                    )
                                    
                                    canopies_data = []  # Initialize canopies list outside the loop
                                    for canopy_idx in range(num_canopies):
                                        st.write(f"Processing canopy {canopy_idx + 1} for {area_name}")
                                        col1, col2 = st.columns(2)
                                        with col1:
                                            reference_number = st.text_input(
                                                "Reference Number",
                                                key=f"ref_{level_idx}_{area_idx}_{canopy_idx}"
                                            )
                                            configuration = st.selectbox(
                                                "Configuration",
                                                options=["WALL", "ISLAND", "OTHER"],
                                                key=f"config_{level_idx}_{area_idx}_{canopy_idx}"
                                            )
                                        # Add fire suppression radio button right after configuration
                                        fire_suppression = st.radio(
                                            "Include Fire Suppression",
                                            options=["No", "Yes"],
                                            key=f"fire_suppression_{level_idx}_{area_idx}_{canopy_idx}"
                                        )
                                    
                                        with col2:
                                            model = st.selectbox(
                                                "Model",
                                                options=["Select..."] + VALID_CANOPY_MODELS,
                                                key=f"model_{level_idx}_{area_idx}_{canopy_idx}"
                                            )
                                            
                                            wall_cladding = st.selectbox(
                                                "Wall Cladding",
                                                options=["Select...", "2M² (HFL)"],
                                                key=f"cladding_{level_idx}_{area_idx}_{canopy_idx}"
                                            )
                                            
                                            if wall_cladding != "Select...":
                                                col1, col2 = st.columns(2)
                                                with col1:
                                                    cladding_width = st.number_input(
                                                        "Cladding Width (mm)",
                                                        min_value=0,
                                                        value=0,
                                                        step=100,
                                                        key=f"cladding_width_{level_idx}_{area_idx}_{canopy_idx}"
                                                    )
                                                with col2:
                                                    cladding_height = st.number_input(
                                                        "Cladding Height (mm)",
                                                        min_value=0,
                                                        value=0,
                                                        step=100,
                                                        key=f"cladding_height_{level_idx}_{area_idx}_{canopy_idx}"
                                                    )
                                                
                                                wall_positions = st.multiselect(
                                                    "Wall Positions",
                                                    options=["Rear", "Left", "Right"],
                                                    key=f"wall_positions_{level_idx}_{area_idx}_{canopy_idx}"
                                                )
                                        
                                        if reference_number:
                                            canopy_data = {
                                                'reference_number': reference_number,
                                                'configuration': configuration,
                                                'model': model,
                                                'fire_suppression': fire_suppression == 'Yes',  # Store as boolean
                                                'level_name': level_name,  # Store level name
                                                'area_name': area_name,    # Store area name
                                                'wall_cladding': {
                                                    'type': wall_cladding,
                                                    'width': cladding_width if wall_cladding != "Select..." else 0,
                                                    'height': cladding_height if wall_cladding != "Select..." else 0,
                                                    'positions': wall_positions if wall_cladding != "Select..." else []
                                                }
                                            }
                                            canopies_data.append(canopy_data)
                                
                                elif cost_sheet == "Reco Air":
                                    # For Reco Air, we only need the RECOAIR system option
                                    include_recoair = "Yes"  # Always include RECOAIR for Reco Air cost sheet
                                    include_uvc = "No"  # No UV-C for Reco Air
                                    include_sdu = "No"  # No SDU for Reco Air
                                    canopies_data = []  # No canopies for Reco Air
                                
                                # Move this outside the canopy loop so we only add the area once with all its canopies
                                # The canopies_data list will be empty if no canopies were added.
                                # An area is added as long as area_name was provided (checked by outer condition).
                                areas_data.append({
                                    'area_name': area_name,
                                    'canopies': canopies_data, 
                                    'include_uvc': include_uvc == "Yes",
                                    'include_recoair': include_recoair == "Yes",
                                    'include_sdu': include_sdu == "Yes"
                                })
                    
                    if areas_data:
                        levels_data.append({
                            'level_name': level_name,
                            'areas': areas_data
                        })
        
        # Update session state with levels data
        st.session_state.project_data['Levels'] = levels_data
        
        # Update sidebar with current data
        create_project_sidebar(st.session_state.project_data)
        
        # Add Installation Location dropdown
        st.markdown("### Delivery Location")
        st.markdown("Select where the equipment will be delivered (this will be written to cell D183 in each sheet):")
        
        # Define delivery locations directly here instead of using dropdowns dictionary
        delivery_locations = [
            "Select...",
            "ABERDEEN 590",
            "ABINGDON 110",
            "ALDEBURGH 112",
            "ALDERSHOT 110",
            "ALNWICK 342",
            "ANDOVER 110",
            "ASHFORD 25",
            "AYLESBURY 86",
            "BANBURY 102",
            "BANGOR 324",
            "BARKING 32",
            "BARNET 55",
            "BARNSLEY 209",
            "BARNSTABLE 227",
            "BARROW-IN-FURNESS 348",
            "BASILDON 38",
            "BASINGSTOKE 82",
            "BATH 154",
            "BEDFORD 103",
            "BERWICK-UPON-TWEED 371",
            "BILLERICAY 37",
            "BIRKENHEAD 277",
            "BIRMINGHAM 168",
            "BLACKBURN 283",
            "BLACKPOOL 289",
            "BLANDFORD FORUM 144",
            "BODMIN 273",
            "BOGNOR REGIS 88",
            "BOLTON 259",
            "BOOTLE 272",
            "BOURNEMOUTH 140",
            "BRADFORD 234",
            "BRAINTREE 60",
            "BRIDGEND 205",
            "BRIDLINGTON 244",
            "BRIGHTON 68",
            "BRISTOL 157",
            "BUCKINGHAMSHIRE 109",
            "BURNLEY 296",
            "BURTON UPON TRENT 175",
            "BURY ST EDMUNDS 98",
            "CAMBRIDGE 85",
            "CANNOCK 175",
            "CANTERBURY 30",
            "CARDIFF 192",
            "CARLISLE 356",
            "CARMARTHEN 252",
            "CHELTENHAM 148",
            "CHESTER 268",
            "COVENTRY 146",
            "CHIPPENHAM 136",
            "COLCHESTER 78",
            "CORBY 128",
            "DARTMOUTH 245",
            "DERBY 178",
            "DONCASTER 203",
            "DORCHESTER 160",
            "DORKING 46",
            "DOVER 45",
            "DURHAM 299",
            "EASTBOURNE 57",
            "EASTLEIGH 109",
            "EDINBURGH 428",
            "ENFIELD 49",
            "EXETER 205",
            "EXMOUTH 207",
            "FELIXSTOWE 103",
            "GATWICK 44",
            "GLASGOW 456",
            "GLASTONBURY 164",
            "GLOUCESTER 151",
            "GRANTHAM 143",
            "GREAT YARMOUTH 147",
            "GRIMSBY 215",
            "GUILDFORD 59",
            "HARLOW 47",
            "HARROGATE 236",
            "HARTLEPOOL 286",
            "HASTINGS 40",
            "HEXHAM 325",
            "HEREFORD 184",
            "HIGH WYCOMBE 80",
            "HIGHBRIDGE 187",
            "HONITON 190",
            "HORSHAM 55",
            "HOUNSLOW 55",
            "HUDDERSFIELD 239",
            "HULL 247",
            "HUNTINGDON 94",
            "INVERNESS 619",
            "IPSWICH 94",
            "IRELAND",
            "KENDAL 321",
            "KETTERING 127",
            "KIDDERMINSTER 179",
            "KILMARNOCK 449",
            "KINGSTON UPON HULL 220",
            "KINGSTON UPON THAMES 52",
            "LANCASTER 290",
            "LAUNCESTON 251",
            "LEAMINGTON SPA 146",
            "LEEDS 231",
            "LEICESTER 151",
            "LEIGH ON SEA 45",
            "LEWISHAM 29",
            "LINCOLN 179",
            "LIVERPOOL 258",
            "LLANDUDNO 309",
            "LONDON in FORS GOLD(varies)",
            "LUTON 80",
            "MABLETHORPE 182",
            "MACCLESFIELD 244",
            "MANCHESTER 251",
            "MARGATE 46",
            "MIDDLESBROUGH 286",
            "MILFORD HAVEN 289",
            "MILTON KEYNES 101",
            "MORPETH 327",
            "NANTWICH 232",
            "NEWBURY 101",
            "NEWCASTLE 308",
            "NEWPORT 178",
            "NEWQUAY 178",
            "NORTHAMPTON 116",
            "NORTHUMBERLAND 341",
            "NORWICH 136",
            "NOTTINGHAM 177",
            "OKEHAMPTON 232",
            "OXFORD 106",
            "PENRITH 316",
            "PENZANCE 318",
            "PERTH 477",
            "PETERBOROUGH 124",
            "PETERSFIELD 87",
            "PETWORTH 71",
            "PLYMOUTH 247",
            "PONTEFRACT 221",
            "POOLE 144",
            "PORTSMOUTH 102",
            "READING 88",
            "REIGATE 39",
            "RINGWOOD 130",
            "ROSS-ON-WYE 171",
            "ROTHERHAM 203",
            "SALISBURY 120",
            "SCARBOROUGH 277",
            "SCUNTHORPE 204",
            "SHEFFIELD 205",
            "SHREWSBURY 207",
            "SHROPSHIRE 218",
            "SLOUGH 72",
            "SOUTH SHIELDS 310",
            "SOUTHAMPTON 112",
            "SOUTHEND 52",
            "SOUTHPORT 279",
            "SPALDING 143",
            "ST ALBANS 62",
            "ST IVES 317",
            "STAFFORD 187",
            "STAINES 61",
            "STEVENAGE 72",
            "STIRLING 445",
            "STOCKPORT 257",
            "STOCKTON 278",
            "STOKE-ON-TRENT 205",
            "STRATFORD UPON AVON 151",
            "SUNDERLAND 309",
            "SWINDON 121",
            "TAMWORTH 180",
            "TAUNTON 185",
            "TELFORD 193",
            "TILBURY 34",
            "TORQUAY 227",
            "TUNBRIDGE WELLS 26",
            "UXBRIDGE 74",
            "WAKEFIELD 214",
            "WARMISTER 137",
            "WARWICK 148",
            "WATFORD 67",
            "WELSHPOOL 238",
            "WEMBLEY 55",
            "WEYMOUTH 173",
            "WHITBY 282",
            "WIGAN 252",
            "WINCANTON 149",
            "WINCHESTER 100",
            "WOKING 60",
            "WOLVERHAMPTON 175",
            "WORCESTER 160",
            "WREXHAM 250",
            "YEOVIL 163",
            "YORK 243"
        ]
        
        delivery_location = st.selectbox(
            "Delivery Location",
            options=delivery_locations,
            key="delivery_location"
        )
        
        if delivery_location != "Select...":
            st.session_state.project_data['Delivery_Location'] = delivery_location
        
        # Add save button and check if all required fields are filled
        required_fields = {
            'Project Name': st.session_state.project_data.get('Project Name', ''),
            'Project Number': st.session_state.project_data.get('Project Number', ''),
            'Customer': st.session_state.project_data.get('Customer', ''),
            'Company': st.session_state.project_data.get('Company', ''),
            'Location': st.session_state.project_data.get('Location', ''),
            'Address': st.session_state.project_data.get('Address', ''),
            'Sales Contact': st.session_state.project_data.get('Sales Contact', ''),
            'Estimator': st.session_state.project_data.get('Estimator', '')
        }
        
        # Check if all required fields are filled and at least one canopy is added
        all_fields_filled = all(value and value != "Select..." for value in required_fields.values())
        has_canopies = any(level.get('areas', []) for level in st.session_state.project_data.get('Levels', []))
        
        st.markdown("---")
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("Save Project", key="save_project_button", disabled=not (all_fields_filled and has_canopies)):
                try:
                    save_to_excel(st.session_state.project_data)
                except Exception as e:
                    st.error(f"Error saving project: {str(e)}")
        
        with col2:
            # Add upload section
            create_upload_section(col2)

def copy_template_sheet(workbook, source_sheet_name, new_sheet_name):
    """Copy a sheet and rename it"""
    source = workbook[source_sheet_name]
    target = workbook.copy_worksheet(source)
    target.title = new_sheet_name
    return target

def get_initials(name):
    """Extract initials from a name, skipping 'Select...'"""
    if name == "Select...":
        return ""
    # Split the name and take first letter of each part
    words = name.split()
    initials = ''.join(word[0].upper() for word in words if word)
    return initials

def write_to_sheet(sheet, data, level_name, area_name, canopies, fs_sheet=None, is_edge_box=False):
    """Write data to specific cells in the sheet"""
    # Write sheet title
    sheet_title = f"{level_name} - {area_name}"
    if is_edge_box:
        # Use cell() method to handle potential merged cells
        try:
            sheet.cell(row=1, column=3, value=sheet_title)  # C1
            
            # For RECOAIR sheets, write item number 1.01 to C12
            if 'RECOAIR' in sheet.title:
                sheet.cell(row=12, column=3, value="1.01")  # C12 - item number for the entire unit
            # For EBOX sheets, write UV-C
            else:
                sheet.cell(row=12, column=3, value="UV-C")  # C12
            
            # Write general info using cell() method
            sheet.cell(row=3, column=4, value=data['Project Number'])  # D3
            customer_company = f"{data['Customer']} ({data['Company']})" if data['Company'] else data['Customer']
            sheet.cell(row=5, column=4, value=customer_company)  # D5
            sheet.cell(row=7, column=4, value=f"{get_initials(data['Sales Contact'])}/{get_initials(data['Estimator'])}")  # D7
            sheet.cell(row=3, column=8, value=data['Project Name'])  # H3
            sheet.cell(row=5, column=8, value=data['Location'])  # H5
            sheet.cell(row=7, column=8, value=data['Date'])  # H7
        except Exception as e:
            st.warning(f"Could not set some cell values: {str(e)}")
        
        # Write the delivery location to appropriate cell based on sheet type
        if 'Delivery_Location' in data and data['Delivery_Location']:
            try:
                # Check if this is an EBOX or RECOAIR sheet
                if 'RECOAIR' in sheet.title:
                    sheet.cell(row=37, column=5, value=data['Delivery_Location'])  # E37
                elif 'EBOX' in sheet.title or 'EDGE BOX' in sheet.title:
                    sheet.cell(row=38, column=5, value=data['Delivery_Location'])  # E38
                else:
                    sheet.cell(row=183, column=4, value=data['Delivery_Location'])  # D183
            except Exception as e:
                st.warning(f"Could not set delivery location: {str(e)}")
            
        return  # Don't write any other data to EDGE BOX sheets
    else:
        # Use cell() method for regular sheets too
        try:
            sheet.cell(row=1, column=2, value=sheet_title)  # B1
        except Exception as e:
            st.warning(f"Could not set sheet title: {str(e)}")
    
    # Write general info to sheets
    def write_general_info(target_sheet, is_edge_box=False):
        # Skip writing to EDGE BOX sheets
        if is_edge_box:
            return
    
        # For regular sheets, use cell() method to handle potential merged cells
        try:
            target_sheet.cell(row=3, column=3, value=data['Project Number'])  # C3
            customer_company = f"{data['Customer']} ({data['Company']})" if data['Company'] else data['Customer']
            target_sheet.cell(row=5, column=3, value=customer_company)  # C5
            target_sheet.cell(row=7, column=3, value=f"{get_initials(data['Sales Contact'])}/{get_initials(data['Estimator'])}")  # C7
            target_sheet.cell(row=3, column=7, value=data['Project Name'])  # G3
            target_sheet.cell(row=5, column=7, value=data['Location'])  # G5
            target_sheet.cell(row=7, column=7, value=data['Date'])  # G7
            target_sheet.cell(row=7, column=15, value='A')  # O7 - Initial revision
        except Exception as e:
            st.warning(f"Could not set some general info values: {str(e)}")
        
        # Write the delivery location to E183 for EBOX and RECOAIR sheets
        if 'Delivery_Location' in data and data['Delivery_Location']:
            try:
                # Check if this is an EBOX or RECOAIR sheet
                if 'RECOAIR' in target_sheet.title:
                    target_sheet.cell(row=37, column=5, value=data['Delivery_Location'])  # E37
                elif 'EBOX' in target_sheet.title or 'EDGE BOX' in target_sheet.title:
                    target_sheet.cell(row=38, column=5, value=data['Delivery_Location'])  # E38
                else:
                    target_sheet.cell(row=183, column=4, value=data['Delivery_Location'])  # D183
            except Exception as e:
                st.warning(f"Could not set delivery location: {str(e)}")
    
    write_general_info(sheet, is_edge_box)
    
    # Handle fire suppression sheet if needed
    if fs_sheet:
        # Write the same title as the canopy sheet, using cell() method
        try:
            fs_sheet.cell(row=1, column=2, value=sheet_title)  # B1
        except Exception as e:
            st.warning(f"Could not set fire suppression sheet title: {str(e)}")
        
        # Get the sheet number from the canopy sheet name
        sheet_number = sheet.title.split('(')[-1].strip(')')
        
        # Rename fire suppression sheet to match canopy sheet format
        fs_sheet.title = f"FIRE SUPP - {level_name} ({sheet_number})"
        
        # Write general info
        write_general_info(fs_sheet)
        
        # Write fire suppression canopies starting from first row
        fs_row = 12  # Start at first row
        for canopy in canopies:
            if canopy.get('fire_suppression', False):
                try:
                    fs_sheet.cell(row=fs_row, column=2, value=canopy['reference_number'])  # B{fs_row}
                except Exception as e:
                    st.warning(f"Could not set fire suppression reference number: {str(e)}")
                fs_row += 17  # Move to next row
                
    # Write canopy data
    for idx, canopy in enumerate(canopies):
        base_row = 12 + (idx * 17)
        
        # Write to sheet using cell() method to handle potential merged cells
        try:
            if canopy['reference_number'] != "Select...":
                sheet.cell(row=base_row, column=2, value=canopy['reference_number'])  # B{base_row}
            if canopy['configuration'] != "Select...":
                sheet.cell(row=base_row + 2, column=3, value=canopy['configuration'])  # C{base_row + 2}
            if canopy['model'] != "Select...":
                sheet.cell(row=base_row + 2, column=4, value=canopy['model'])  # D{base_row + 2}
            
            # Write standard entries
            sheet.cell(row=base_row + 3, column=3, value="LIGHT SELECTION")  # C{base_row + 3}
            sheet.cell(row=base_row + 4, column=3, value="SELECT WORKS")  # C{base_row + 4}
            sheet.cell(row=base_row + 5, column=3, value="SELECT WORKS")  # C{base_row + 5}
            sheet.cell(row=base_row + 6, column=3, value="BIM/ REVIT per CANOPY")  # C{base_row + 6}
            sheet.cell(row=base_row + 6, column=4, value="1")  # D{base_row + 6}
            
            # Write dimensions if available
            if 'length' in canopy and canopy['length']:
                sheet.cell(row=base_row + 2, column=5, value=canopy['length'])  # E{base_row + 2}
            if 'width' in canopy and canopy['width']:
                sheet.cell(row=base_row + 2, column=6, value=canopy['width'])  # F{base_row + 2}
            if 'height' in canopy and canopy['height']:
                sheet.cell(row=base_row + 2, column=7, value=canopy['height'])  # G{base_row + 2}
                
            # Write base price if available
            if 'base_price' in canopy and canopy['base_price']:
                sheet.cell(row=base_row + 2, column=11, value=canopy['base_price'])  # K{base_row + 2}
                
            # Write K9 price if available
            if 'k9_price' in canopy and canopy['k9_price']:
                sheet.cell(row=base_row + 2, column=19, value=canopy['k9_price'])  # S{base_row + 2}
            
            # Wall cladding - only write if not "Select..."
            if canopy['wall_cladding']['type'] and canopy['wall_cladding']['type'] != "Select...":
                cladding_row = base_row + 7
                # Write wall cladding type to C column
                sheet.cell(row=cladding_row, column=3, value=canopy['wall_cladding']['type'])  # C{cladding_row}
                
                # Write dimensions and positions to hidden cells for storage
                sheet.cell(row=cladding_row, column=17, value=canopy['wall_cladding']['width'])  # Q{cladding_row}
                sheet.cell(row=cladding_row, column=18, value=canopy['wall_cladding']['height'])  # R{cladding_row}
                sheet.cell(row=cladding_row, column=19, value=','.join(canopy['wall_cladding']['positions']))  # S{cladding_row}
                sheet.cell(row=cladding_row, column=20, value=f"2M² (HFL) - {canopy['wall_cladding']['width']}x{canopy['wall_cladding']['height']}mm ({'/'.join(canopy['wall_cladding']['positions'])})") # T{cladding_row}
            
            # Write MUA VOL to Q22 if it's an F-type canopy
            model = str(canopy.get('model', ''))  # Convert to string first
            if 'F' in model.upper():
                print(f"Writing MUA VOL for {model}: {canopy.get('mua_vol', '-')}")
                sheet.cell(row=22 + (idx * 17), column=17, value=f"MUA VOL: {canopy.get('mua_vol', '-')}")  # Q{22 + (idx * 17)}
        except Exception as e:
            st.warning(f"Could not set some canopy values: {str(e)}")
    

def add_dropdowns_to_sheet(wb, sheet, start_row):
    """Add data validation (dropdowns) to specific cells"""
    if 'Lists' not in wb.sheetnames:
        list_sheet = wb.create_sheet('Lists')
    else:
        list_sheet = wb['Lists']

    # Define all dropdown options
    dropdowns = {
        'lights': {
            'options': [
                'LED STRIP L6 Inc DALI',
                'LED STRIP L12 inc DALI', 
                'LED STRIP L18 Inc DALI',
                'Small LED Spots inc DALI',
                'LARGE LED Spots inc DALI'
            ],
            'column': 'A',
            'target_col': 'C',
            'row_offset': 3  # C15 for first canopy
        },
        'special_works_1': {
            'options': [
                'ROUND CORNERS',
                'CUT OUT',
                'CASTELLE LOCKING ',
                'HEADER DUCT S/S',
                'HEADER DUCT',
                'PAINT FINSH',
                'UV ON DEMAND',
                'E/over for emergency strip light',
                'E/over for small emer. spot light',
                'E/over for large emer. spot light',
                'COLD MIST ON DEMAND',
                'CMW  PIPEWORK HWS/CWS',
                'CANOPY GROUND SUPPORT',
                '2nd EXTRACT PLENUM',
                'SUPPLY AIR PLENUM',
                'CAPTUREJET PLENUM',
                'COALESCER'
            ],
            'column': 'B',
            'target_col': 'C',
            'row_offset': 4  # C16 for first special work
        },
        'special_works_2': {
            'options': [
                'SELECT WORKS',  # Add default option
                'ROUND CORNERS',
                'CUT OUT',
                'CASTELLE LOCKING',
                'HEADER DUCT S/S',
                'HEADER DUCT',
                'PAINT FINSH',
                'UV ON DEMAND',
                'E/over for emergency strip light',
                'E/over for small emer. spot light',
                'E/over for large emer. spot light',
                'COLD MIST ON DEMAND',
                'CMW  PIPEWORK HWS/CWS',
                'CANOPY GROUND SUPPORT',
                '2nd EXTRACT PLENUM',
                'SUPPLY AIR PLENUM',
                'CAPTUREJET PLENUM',
                'COALESCER'
            ],
            'column': 'B',  # Use same column as they're the same options
            'target_col': 'C',
            'row_offset': 5  # C17 for second special work
        },
        'wall_cladding': {
            'options': ['', '2M² (HFL)'],
            'column': 'C',
            'target_col': 'C',
            'row_offset': 7  # C19 for first canopy
        },
        'control_panel': {
            'options': ['CP1S', 'CP2S', 'CP3S', 'CP4S'],
            'column': 'D',
            'target_col': 'C',
            'row_offset': 13  # C25 for first canopy
        },
        'ww_pods': {
            'options': [
                '1000-S', '1500-S', '2000-S', '2500-S', '3000-S',
                '1000-D', '1500-D', '2000-D', '2500-D', '3000-D'
            ],
            'column': 'E',
            'target_col': 'C',
            'row_offset': 14  # C26 for first canopy
        },
        'delivery_location': {
            'options': [
                "",  # Empty option first
                "ABERDEEN 590",
                "ABINGDON 110",
                "ALDEBURGH 112",
                "ALDERSHOT 110",
                "ALNWICK 342",
                "ANDOVER 110",
                "ASHFORD 25",
                "AYLESBURY 86",
                "BANBURY 102",
                "BANGOR 324",
                "BARKING 32",
                "BARNET 55",
                "BARNSLEY 209",
                "BARNSTABLE 227",
                "BARROW-IN-FURNESS 348",
                "BASILDON 38",
                "BASINGSTOKE 82",
                "BATH 154",
                "BEDFORD 103",
                "BERWICK-UPON-TWEED 371",
                "BILLERICAY 37",
                "BIRKENHEAD 277",
                "BIRMINGHAM 168",
                "BLACKBURN 283",
                "BLACKPOOL 289",
                "BLANDFORD FORUM 144",
                "BODMIN 273",
                "BOGNOR REGIS 88",
                "BOLTON 259",
                "BOOTLE 272",
                "BOURNEMOUTH 140",
                "BRADFORD 234",
                "BRAINTREE 60",
                "BRIDGEND 205",
                "BRIDLINGTON 244",
                "BRIGHTON 68",
                "BRISTOL 157",
                "BUCKINGHAMSHIRE 109",
                "BURNLEY 296",
                "BURTON UPON TRENT 175",
                "BURY ST EDMUNDS 98",
                "CAMBRIDGE 85",
                "CANNOCK 175",
                "CANTERBURY 30",
                "CARDIFF 192",
                "CARLISLE 356",
                "CARMARTHEN 252",
                "CHELTENHAM 148",
                "CHESTER 268",
                "COVENTRY 146",
                "CHIPPENHAM 136",
                "COLCHESTER 78",
                "CORBY 128",
                "DARTMOUTH 245",
                "DERBY 178",
                "DONCASTER 203",
                "DORCHESTER 160",
                "DORKING 46",
                "DOVER 45",
                "DURHAM 299",
                "EASTBOURNE 57",
                "EASTLEIGH 109",
                "EDINBURGH 428",
                "ENFIELD 49",
                "EXETER 205",
                "EXMOUTH 207",
                "FELIXSTOWE 103",
                "GATWICK 44",
                "GLASGOW 456",
                "GLASTONBURY 164",
                "GLOUCESTER 151",
                "GRANTHAM 143",
                "GREAT YARMOUTH 147",
                "GRIMSBY 215",
                "GUILDFORD 59",
                "HARLOW 47",
                "HARROGATE 236",
                "HARTLEPOOL 286",
                "HASTINGS 40",
                "HEXHAM 325",
                "HEREFORD 184",
                "HIGH WYCOMBE 80",
                "HIGHBRIDGE 187",
                "HONITON 190",
                "HORSHAM 55",
                "HOUNSLOW 55",
                "HUDDERSFIELD 239",
                "HULL 247",
                "HUNTINGDON 94",
                "INVERNESS 619",
                "IPSWICH 94",
                "IRELAND",
                "KENDAL 321",
                "KETTERING 127",
                "KIDDERMINSTER 179",
                "KILMARNOCK 449",
                "KINGSTON UPON HULL 220",
                "KINGSTON UPON THAMES 52",
                "LANCASTER 290",
                "LAUNCESTON 251",
                "LEAMINGTON SPA 146",
                "LEEDS 231",
                "LEICESTER 151",
                "LEIGH ON SEA 45",
                "LEWISHAM 29",
                "LINCOLN 179",
                "LIVERPOOL 258",
                "LLANDUDNO 309",
                "LONDON in FORS GOLD(varies)",
                "LUTON 80",
                "MABLETHORPE 182",
                "MACCLESFIELD 244",
                "MANCHESTER 251",
                "MARGATE 46",
                "MIDDLESBROUGH 286",
                "MILFORD HAVEN 289",
                "MILTON KEYNES 101",
                "MORPETH 327",
                "NANTWICH 232",
                "NEWBURY 101",
                "NEWCASTLE 308",
                "NEWPORT 178",
                "NEWQUAY 178",
                "NORTHAMPTON 116",
                "NORTHUMBERLAND 341",
                "NORWICH 136",
                "NOTTINGHAM 177",
                "OKEHAMPTON 232",
                "OXFORD 106",
                "PENRITH 316",
                "PENZANCE 318",
                "PERTH 477",
                "PETERBOROUGH 124",
                "PETERSFIELD 87",
                "PETWORTH 71",
                "PLYMOUTH 247",
                "PONTEFRACT 221",
                "POOLE 144",
                "PORTSMOUTH 102",
                "READING 88",
                "REIGATE 39",
                "RINGWOOD 130",
                "ROSS-ON-WYE 171",
                "ROTHERHAM 203",
                "SALISBURY 120",
                "SCARBOROUGH 277",
                "SCUNTHORPE 204",
                "SHEFFIELD 205",
                "SHREWSBURY 207",
                "SHROPSHIRE 218",
                "SLOUGH 72",
                "SOUTH SHIELDS 310",
                "SOUTHAMPTON 112",
                "SOUTHEND 52",
                "SOUTHPORT 279",
                "SPALDING 143",
                "ST ALBANS 62",
                "ST IVES 317",
                "STAFFORD 187",
                "STAINES 61",
                "STEVENAGE 72",
                "STIRLING 445",
                "STOCKPORT 257",
                "STOCKTON 278",
                "STOKE-ON-TRENT 205",
                "STRATFORD UPON AVON 151",
                "SUNDERLAND 309",
                "SWINDON 121",
                "TAMWORTH 180",
                "TAUNTON 185",
                "TELFORD 193",
                "TILBURY 34",
                "TORQUAY 227",
                "TUNBRIDGE WELLS 26",
                "UXBRIDGE 74",
                "WAKEFIELD 214",
                "WARMISTER 137",
                "WARWICK 148",
                "WATFORD 67",
                "WELSHPOOL 238",
                "WEMBLEY 55",
                "WEYMOUTH 173",
                "WHITBY 282",
                "WIGAN 252",
                "WINCANTON 149",
                "WINCHESTER 100",
                "WOKING 60",
                "WOLVERHAMPTON 175",
                "WORCESTER 160",
                "WREXHAM 250",
                "YEOVIL 163",
                "YORK 243"
            ],
            'column': 'F',
            'target_col': 'D',
            'target_row': 183  # Fixed row for delivery location
        },
        'plant_hire_1': {
            'options': [
                "", "SL10 GENIE", "EXTENSION FORKS", "2.5M COMBI LADDER",
                "1.5M PODIUM", "3M TOWER", "COMBI LADDER", "PECO LIFT",
                "3M YOUNGMAN BOARD", "GS1930 SCISSOR LIFT",
                "4-6 SHERASCOPIC", "7-9 SHERASCOPIC"
            ],
            'column': 'G',
            'target_col': 'D',
            'target_row': 184  # Fixed row for first plant hire
        },
        'plant_hire_2': {
            'options': [
                "", "SL10 GENIE", "EXTENSION FORKS", "2.5M COMBI LADDER",
                "1.5M PODIUM", "3M TOWER", "COMBI LADDER", "PECO LIFT",
                "3M YOUNGMAN BOARD", "GS1930 SCISSOR LIFT",
                "4-6 SHERASCOPIC", "7-9 SHERASCOPIC"
            ],
            'column': 'G',  # Can use same column as plant_hire_1
            'target_col': 'D',
            'target_row': 185  # Fixed row for second plant hire
        }
    }

    # Write options to Lists sheet and create validations
    for name, config in dropdowns.items():
        # Write options to Lists sheet
        for i, option in enumerate(config['options'], 1):
            list_sheet[f"{config['column']}{i}"] = option
        
        # Create range reference
        range_ref = f"Lists!${config['column']}$1:${config['column']}${len(config['options'])}"
        
        # Create validation
        dv = DataValidation(
            type="list",
            formula1=range_ref,
            allow_blank=True
        )
        sheet.add_data_validation(dv)
        
        if 'target_row' in config:
            # Fixed position dropdown (delivery/installation)
            dv.add(f"{config['target_col']}{config['target_row']}")
        else:
            # Repeating dropdowns (canopy-related)
            current_row = start_row + config['row_offset']
            while current_row <= sheet.max_row:
                dv.add(f"{config['target_col']}{current_row}")
                current_row += 17

def get_canopy_description(model, count):
    """Generate appropriate description based on canopy model"""
    model = model.upper()
    count_str = f"{count}no"  # Format the count properly
    
    # CMWF/CMWI type canopies (Water Wash)
    if model.startswith('CMW'):
        if 'F' in model:
            return f"{count_str} Extract/Supply Canopy c/w Capture Jet Tech and Water Wash Function"
        else:
            return f"{count_str} Extract Canopy c/w Capture Jet Tech and Water Wash Function"
    
    # UV type canopies
    elif model.startswith('UV'):
        if 'F' in model:
            return f"{count_str} Extract/Supply Canopies c/w Capture Jet Tech and UV-c Filtration"
        else:
            return f"{count_str} Extract Canopies c/w Capture Jet Tech and UV-c Filtration"
    
    # CXW type canopies
    elif model.startswith('CXW'):
        return f"{count_str} Condense Canopies c/w Extract and Led Light"
    
    # Standard canopies (KV, etc)
    else:
        if 'F' in model:
            return f"{count_str} Extract/Supply Canopies c/w Capture Jet Tech"
        else:
            return f"{count_str} Extract Canopies c/w Capture Jet Tech"

def format_price(value, return_numeric=False):
    """Format price with commas and 2 decimal places
    
    Args:
        value: Value to format
        return_numeric: If True, returns tuple of (formatted_string, numeric_value)
                       If False, returns only the formatted string
    """
    try:
        # Convert to float and round up
        num = math.ceil(float(str(value).replace(',', '')))
        # Format with commas and 2 decimal places
        formatted = f"{num:,.2f}"
        return (formatted, num) if return_numeric else formatted
    except (ValueError, TypeError):
        return ("0.00", 0) if return_numeric else "0.00"

def write_address_to_lists(workbook, address):
    """Write address to LISTS sheet"""
    try:
        if 'Lists' in workbook.sheetnames:
            lists_sheet = workbook['Lists']
            lists_sheet['L1'] = address
            print(f"Successfully wrote address to Lists sheet L1: {address}")
        elif 'LISTS' in workbook.sheetnames:
            lists_sheet = workbook['LISTS']
            lists_sheet['L1'] = address
            print(f"Successfully wrote address to LISTS sheet L1: {address}")
        else:
            print("Warning: Lists sheet not found. Creating new sheet.")
            lists_sheet = workbook.create_sheet('Lists')
            lists_sheet['L1'] = address
    except Exception as e:
        print(f"Warning: Could not write address to Lists sheet: {str(e)}")

def read_address_from_lists(workbook):
    """Read address from LISTS sheet"""
    try:
        if 'Lists' in workbook.sheetnames:
            lists_sheet = workbook['Lists']
            return lists_sheet['L1'].value or ''
        elif 'LISTS' in workbook.sheetnames:
            lists_sheet = workbook['LISTS']
            return lists_sheet['L1'].value or ''
        else:
            return ''
    except Exception as e:
        print(f"Warning: Could not read address from Lists sheet: {str(e)}")
        return ''

def write_to_word_doc(data, project_data, output_path="output.docx"):
    """Write project information to a Word document using Jinja template"""
    try:
        f_canopy_mua_vols = []
        ww_canopies = []
        has_water_wash_canopies = False
        has_any_uv_canopy = False  # Rename to clarify this is global flag
        has_sdu = False  # Add flag for SDU
        
        # Get water wash canopies from project_data if available
        if project_data and 'water_wash_canopies' in project_data and project_data['water_wash_canopies']:
            ww_canopies = project_data['water_wash_canopies']
            has_water_wash_canopies = True
        
        # Also check for water_wash_items if available
        if not ww_canopies and project_data and 'water_wash_items' in project_data and project_data['water_wash_items']:
            ww_canopies = project_data['water_wash_items'] 
            has_water_wash_canopies = True
        
        # Check for SDU in project_data
        if project_data and 'has_sdu' in project_data:
            has_sdu = project_data['has_sdu']
            
        # Also explicitly check for any SDU sheets
        if project_data and 'sheets' in project_data:
            for sheet in project_data['sheets']:
                sheet_name = sheet.get('sheet_name', '')
                if 'SDU' in sheet_name:
                    has_sdu = True
                    break
                    
                # Also check the title field if available
                sheet_title = sheet.get('title', '')
                if 'SDU' in sheet_title:
                    has_sdu = True
                    break
        
        # Process sheets and canopies
        if project_data and 'sheets' in project_data and project_data['sheets']:
            for sheet in project_data['sheets']:
                if 'canopies' in sheet:
                    for canopy in sheet['canopies']:
                        # Skip non-canopy rows
                        if (canopy.get('reference_number') == 'ITEM' or 
                            canopy.get('model') == 'CANOPY TYPE' or
                            canopy.get('reference_number') == 'DELIVERY & INSTALLATION' or
                            not any(valid_model in str(canopy.get('model', '')).upper() 
                                for valid_model in VALID_CANOPY_MODELS)):  # Check against valid models
                            continue
                        
                        # Convert model and configuration to strings and handle None values
                        model = str(canopy.get('model', '')) if canopy.get('model') is not None else ''
                        config = str(canopy.get('configuration', '')) if canopy.get('configuration') is not None else ''
                        
                        # Check for UV canopies
                        if ('UV' in model.upper() or 'UV' in config.upper()):
                            has_any_uv_canopy = True
                        
                        # Check for water wash canopies if we don't have them from water_wash_items
                        if ('CMWI' in model.upper() or 'CMWF' in model.upper()) and not ww_canopies:
                            has_water_wash_canopies = True
                            
                            # Only append if not already loaded from water_wash_items
                            # Create a comprehensive canopy object with water wash data
                            water_wash_canopy = {
                                'reference_number': canopy.get('reference_number', ''),
                                'model': model,
                                'cws_2bar': canopy.get('cws_2bar', '-'),
                                'hws_2bar': canopy.get('hws_2bar', '-'),
                                'hws_storage': canopy.get('hws_storage', '-'),
                                'ww_price': canopy.get('ww_price', 0),
                                'ww_control_price': canopy.get('ww_control_price', 0),
                                'ww_install_price': canopy.get('ww_install_price', 0),
                                'control_panel': canopy.get('control_panel', '-'),
                                'ww_pods': canopy.get('ww_pods', '-'),
                                'ww_control': canopy.get('ww_control', '-')
                            }
                            ww_canopies.append(water_wash_canopy)
                            # st.write(f"DEBUG: Added water wash canopy from sheets: {canopy.get('reference_number')}")
                        
                        # Check for F-type canopies
                        if 'F' in model.upper() and canopy.get('mua_vol'):
                            f_canopy_mua_vols.append({
                                'item_no': canopy['reference_number'],
                                'model': model,
                                'mua_vol': canopy['mua_vol']
                            })
        
        # Get address safely from project_data - various fallback methods
        address = ''
        # Method 1: Try to get from project_data.sheets[0].project_info.address
        if project_data and 'sheets' in project_data and project_data['sheets']:
            first_sheet = project_data['sheets'][0]
            if 'project_info' in first_sheet:
                address = first_sheet['project_info'].get('address', '')
                
        # Method 2: Try to get from data directly (if passed from form)
        if not address and data and 'Address' in data:
            address = data.get('Address', '')
        
        # Method 3: If form data has Company, look up in company_addresses dictionary
        if not address and data and 'Company' in data and data['Company'] in company_addresses:
            address = company_addresses[data['Company']]  # company_addresses directly contains address strings
        
        # Load the template
        doc = DocxTemplate("resources/Halton Quote Feb 2024 (1).docx")
        
        # Get the full sales contact name from initials
        sales_initials = data['Sales Contact']
        sales_contact_name = next((name for name in contacts.keys() if name.startswith(sales_initials)), None)
        
        # If we couldn't find the full name, try to get it from the contacts dictionary
        if not sales_contact_name:
            # Try to find a matching contact by initials
            for full_name in contacts.keys():
                if get_initials(full_name) == sales_initials:
                    sales_contact_name = full_name
                    break
        
        # If we still don't have a name, use the initials
        if not sales_contact_name:
            st.error(f"Could not find contact information for {sales_initials}")
            sales_contact_name = sales_initials
            contact_number = ""
        else:
            contact_number = contacts[sales_contact_name]
        
        # Extract customer's first name
        customer_full = data['Customer']
        customer_first_name = customer_full.split()[0] if customer_full else ""
        
        # Get estimator initials from the sales_estimator field (format: "sales/estimator")
        estimator_initials = data.get('Estimator', '').split('/')[-1] if '/' in data.get('Estimator', '') else ''
        
        # Find estimator name and role from initials
        estimator_name = ''
        estimator_role = ''
        for full_name, role in estimators.items():
            if get_initials(full_name) == estimator_initials:
                estimator_name = full_name
                estimator_role = role
                break
        
        # Get current revision
        current_revision = project_data['sheets'][0].get('revision', 'A')
        
        # Create Halton reference with project number, initials and revision
        halton_ref = f"{data['Project Number']}/{get_initials(sales_contact_name)}/{estimator_initials}/{current_revision}" if current_revision != 'A' else f"{data['Project Number']}/{get_initials(sales_contact_name)}/{estimator_initials}"
        
        # Create quote title based on revision
        quote_title = "QUOTATION" if current_revision == 'A' else f"QUOTATION - Revision {current_revision}"
        
        # Initialize global flags and arrays
        has_water_wash_canopies = False
        
        # Collect all cladding items from the project_data structure
        cladding_items = []
        
        # Check if wall_cladding is in project_data (for backward compatibility)
        if 'wall_cladding' in project_data:
            # Collect all cladding items from all areas
            for area_name, area_cladding in project_data['wall_cladding'].items():
                cladding_items.extend(area_cladding)
        else:
            # Fallback to old method of collecting cladding from canopies
            for sheet in project_data['sheets']:
                for canopy in sheet['canopies']:
                    # Skip non-canopy rows
                    if (canopy['reference_number'] == 'ITEM' or 
                        canopy['model'] == 'CANOPY TYPE' or
                        canopy['reference_number'] == 'DELIVERY & INSTALLATION' or
                        not any(valid_model in str(canopy['model']).upper() 
                               for valid_model in VALID_CANOPY_MODELS)):  # Check against valid models
                        continue
                        
                    # Convert model and configuration to strings
                    model = str(canopy.get('model', '')) if canopy.get('model') is not None else ''
                    config = str(canopy.get('configuration', '')).upper() if canopy.get('configuration') is not None else ''
                    
                    # Check for UV canopies
                    if ('UV' in model.upper() or 'UV' in config):
                        has_any_uv_canopy = True
                    
                    # Check for water wash canopies
                    if 'CMWI' in model.upper() or 'CMWF' in model.upper():
                        has_water_wash_canopies = True
                        ww_canopies.append(canopy)
                    
                    # Check for F-type canopies
                    if 'F' in model.upper() and canopy.get('mua_vol'):
                        f_canopy_mua_vols.append({
                            'item_no': canopy['reference_number'],
                            'model': model,
                            'mua_vol': canopy['mua_vol']
                        })
                    
                    # Create description based on selected positions
                    positions = canopy['wall_cladding']['positions']
                    position_str = '/'.join(positions).lower() if positions else ''
                    
                    if position_str:  # Only add if there are positions selected
                        cladding_items.append({
                            'item_no': canopy['reference_number'],
                            'description': f"Cladding below Item {canopy['reference_number']}, supplied and installed",
                            'width': canopy['wall_cladding']['width'],
                            'height': canopy['wall_cladding']['height'],
                            'positions': canopy['wall_cladding']['positions'],
                            'price': format_price(math.ceil(float(canopy['wall_cladding'].get('price', 0))))  # Keep as numeric value for calculations
                            
                        })

        # Prepare areas data with technical specifications
        areas = []
        processed_area_names = set()  # Keep track of processed areas to prevent duplication
        
        for sheet in project_data['sheets']:
            display_name = sheet['sheet_name']
            
            # Skip if we've already processed this area
            if display_name in processed_area_names:
                continue
            processed_area_names.add(display_name)
            
            # Get UV price if available
            uv_price = 0
            uv_cost = 0
            total_price = 0
            
            # Check if this specific area has UV canopy by checking uv_control_data
            has_uv_canopy = display_name in project_data['uv_control_data']
            
            if has_uv_canopy:
                uv_data = project_data['uv_control_data'].get(display_name, {})
                uv_price = uv_data.get('n9_price', 0)
                uv_cost = uv_data.get('k9_price', 0)
                total_price = uv_data.get('total_price', 0)  # Get total price from UV control data
                #st.write(f"Debug: Found UV price for {display_name}: {project_data['uv_control_data']}")
               # st.write(f"Debug: Found UV price: {uv_price}")
            
            # Check if this specific area has RECOAIR
            has_recoair = display_name in project_data['recoair_control_data']
            
            # Get RECOAIR price if available
            recoair_price = 0
            recoair_cost = 0
            recoair_total_price = 0
            
            if has_recoair:
                recoair_data = project_data['recoair_control_data'].get(display_name, {})
                recoair_price = recoair_data.get('n9_price', 0)
                recoair_cost = recoair_data.get('k9_price', 0)
                recoair_total_price = recoair_data.get('total_price', 0)
               # st.write(f"Debug: Found RECOAIR price for {display_name}: {project_data['recoair_control_data']}")
               # st.write(f"Debug: Found RECOAIR price: {recoair_price}")
            
            # Add area data to context
            area_data = {
                'name': display_name,
                'has_uv_canopy': has_uv_canopy,  # This now correctly reflects if this specific area has UV
                'uv_price': uv_price,
                'uv_cost': uv_cost,
                'total_price': total_price,  # Add total price to area data
                'has_recoair': has_recoair,  # Add flag for RECOAIR
                'recoair_price': recoair_price,
                'recoair_cost': recoair_cost,
                'recoair_total_price': recoair_total_price,
                'canopies': []
            }
            
            canopies = []
            has_fire_suppression = False
            fire_suppression_canopies = []
            fs_canopy_count = 0
            per_canopy_delivery = 0  # Initialize the per-canopy delivery share
            
            # Get MUA calculations for this area
            mua_calcs = sheet.get('mua_calculations', {})
            total_extract = mua_calcs.get('total_extract_volume', 0)
            required_mua = mua_calcs.get('required_mua', 0)
            total_mua = mua_calcs.get('total_mua_volume', 0)
            mua_shortfall = mua_calcs.get('mua_shortfall', 0)
            
            # Format the important note for this area
            important_note = (
                f"Important Note: - The make-up air flows shown above are the maximum that we can introduce through the "
                f"canopy. This should be equal to approximately 85% of the extract i.e. {required_mua}m³/s\n"
                f"In this instance it only totals {total_mua}m³/s therefore the shortfall of "
                f"{mua_shortfall}m³/s must be introduced through ceiling grilles or diffusers, by others.\n"
                f"If you require further guidance on this, please do not hesitate to contact us."
            ) if total_extract > 0 else ""
            
            # Process canopies and check for UV
            for canopy in sheet['canopies']:
                if (canopy.get('reference_number') and 
                    canopy['reference_number'] != 'ITEM' and 
                    canopy.get('model') and 
                    canopy['model'] != 'CANOPY TYPE' and
                    canopy['reference_number'] != 'DELIVERY & INSTALLATION'):
                    
                    # Convert model and configuration to strings
                    model = str(canopy.get('model', '')) if canopy.get('model') is not None else ''
                    config = str(canopy.get('configuration', '')) if canopy.get('configuration') is not None else ''
                    
                    # Check for UV canopies
                    if ('UV' in model.upper() or 'UV' in config.upper()):
                        has_any_uv_canopy = True
                    
                    # Check for fire suppression
                    if canopy.get('has_fire_suppression') or canopy.get('fire_suppression_data') is not None:
                        has_fire_suppression = True
                        fire_suppression_canopies.append(canopy)
                        fs_canopy_count += 1
                        print(f"DEBUG: Found fire suppression canopy: {canopy.get('reference_number')}")
                    
                    canopies.append({
                        'item_no': canopy['reference_number'],  # Ensure item_no is included
                        'reference_number': canopy['reference_number'],
                        'model': model,
                        'configuration': config,
                        'length': canopy.get('length', '-'),
                        'width': canopy.get('width', '-'),
                        'height': canopy.get('height', '-'),
                        'sections': canopy.get('sections', '-'),
                        'ext_vol': canopy.get('ext_vol', '-'),
                        'ext_static': canopy.get('ext_static', '-'),
                        'mua_vol': canopy.get('mua_vol', '-'),
                        'supply_static': 45 if 'F' in str(canopy.get('model', '')).upper() else '-',
                        'lighting': canopy.get('lighting', '-'),
                        'base_price': format_price(canopy.get('base_price', 0)),
                        'has_fire_suppression': canopy.get('has_fire_suppression', False),
                        'fire_suppression_data': canopy.get('fire_suppression_data', None)
                    })
            
            # Calculate fire suppression totals
            fs_total = 0
            fs_canopies_total = 0
            if has_fire_suppression:
                fs_base_total = sum(safe_convert_to_float(fs_canopy.get('fire_suppression_data', {}).get('base_price', 0))
                                  for fs_canopy in fire_suppression_canopies)
                fs_install_price = safe_convert_to_float(sheet.get('fire_suppression_install', 0))
                
                # Calculate per-canopy installation share
                fs_canopy_count = len(fire_suppression_canopies)
                per_canopy_install_share = fs_install_price / fs_canopy_count if fs_canopy_count > 0 else 0
                
                # Add debug prints to diagnose the value of per_canopy_install_share
                print(f"DEBUG: fs_install_price = {fs_install_price}")
                print(f"DEBUG: fs_canopy_count = {fs_canopy_count}")
                print(f"DEBUG: per_canopy_install_share = {per_canopy_install_share}")
                
                fs_total = fs_base_total + fs_install_price
                fs_canopies_total = sum(safe_convert_to_float(fs_canopy.get('fire_suppression_data', {}).get('base_price', 0))
                                      for fs_canopy in fire_suppression_canopies) + fs_install_price
            
            # Get UV price from project_data if area has UV canopy
            uv_price = 0
            if has_any_uv_canopy:
                #st.write(f"Debug: Looking for UV price for area {display_name}")
                uv_price = project_data.get('uv_control_data', {}).get(display_name, {}).get('price', 0)
                #st.write(f"Debug: Found UV price: {uv_price}")
            
            # Calculate area totals with rounding at each step
            canopy_total = sum(safe_convert_to_float(canopy.get('base_price', 0)) for canopy in canopies)
            delivery_total = safe_convert_to_float(sheet.get('delivery_install', 0))
            commissioning_total = safe_convert_to_float(sheet.get('commissioning_price', 0))
            
            # Calculate total for this area including all components
            area_total = canopy_total + delivery_total + commissioning_total
           # if has_fire_suppression:
               # area_total += fs_total
           # if has_any_uv_canopy:
               # area_total += uv_price
            
            # Get cladding items for this area
            area_cladding_items = []
            
            # Check if we have the new wall_cladding structure in project_data
            if 'wall_cladding' in project_data and display_name in project_data['wall_cladding']:
                # Use the pre-populated area cladding items
                area_cladding_items = project_data['wall_cladding'][display_name]
            else:
                # Fallback to old method of collecting cladding from canopies
                #NOT HERE
                for canopy in canopies:
                    if canopy.get('wall_cladding', {}).get('type') and canopy['wall_cladding']['type'] != "Select...":
                        area_cladding_items.append({
                            'item_no': canopy['reference_number'],
                            'description': f"Cladding below Item {canopy['reference_number']}, supplied and installed",
                            'width': canopy['wall_cladding']['width'],
                            'height': canopy['wall_cladding']['height'],
                            'positions': canopy['wall_cladding']['positions'],
                            'price': safe_convert_to_float(canopy['wall_cladding'].get('price', 0))  # Store as numeric value, not formatted string
                        })
            
            # Calculate sheet total including cladding
            sheet_total = area_total  # Start with area_total which already includes canopies, delivery, commissioning
            if area_cladding_items:
                cladding_total = sum(safe_convert_to_float(item['price']) for item in area_cladding_items)
                sheet_total += cladding_total
            
            # Add fire suppression total from canopies
            if has_fire_suppression:
                fs_canopies_total = sum(
                    safe_convert_to_float(fs_canopy.get('fire_suppression_data', {}).get('base_price', 0)) +
                    safe_convert_to_float(fs_canopy.get('fire_suppression_data', {}).get('install_price', 0))
                    for fs_canopy in fire_suppression_canopies
                )
                sheet_total += fs_canopies_total
            
            # Add UV-C price if not already included
            if display_name in project_data['uv_control_data']:
                sheet_total += project_data['uv_control_data'].get(display_name, {}).get('total_price', 0)
            
            # RECOAIR price is intentionally excluded from area totals per request
            # if display_name in project_data['recoair_control_data']:
            #     sheet_total += project_data['recoair_control_data'].get(display_name, {}).get('total_price', 0)
            
            areas.append({
                'name': display_name,
                'canopies': canopies,  # Now includes item_no
                'has_uv': display_name in project_data['uv_control_data'],
                'uv_price': format_price(project_data['uv_control_data'].get(display_name, {}).get('n9_price', 0)) if display_name in project_data['uv_control_data'] else None,
                'has_recoair': display_name in project_data['recoair_control_data'],
                'recoair_price': format_price(project_data['recoair_control_data'].get(display_name, {}).get('n9_price', 0)) if display_name in project_data['recoair_control_data'] else None,
                'has_fire_suppression': has_fire_suppression,
                'fire_suppression_canopies': [{
                    'item_no': fs_canopy['reference_number'],  # Include item_no
                    'reference_number': fs_canopy['reference_number'],  # Keep reference_number too
                    'base_price': format_price(fs_canopy.get('fire_suppression_data', {}).get('base_price', 0)),
                    'install_share': format_price(math.ceil(fs_canopy.get('fire_suppression_data', {}).get('install_price', 0))),
                    'total_price': format_price(
                        safe_convert_to_float(fs_canopy.get('fire_suppression_data', {}).get('base_price', 0)) +
                        safe_convert_to_float(fs_canopy.get('fire_suppression_data', {}).get('install_price', 0))
                    ),
                    'system_description': 'Ansul R 102 System' or fs_canopy.get('fire_suppression_data', {}).get('system_description', ''),
                    'tank_quantity': fs_canopy.get('fire_suppression_data', {}).get('tank_quantity', ''),
                    'manual_release': fs_canopy.get('fire_suppression_data', {}).get('manual_release', '')
                } for fs_canopy in fire_suppression_canopies],
                'fire_suppression_total': format_price(fs_total),
                'fire_suppression_data': {
                    'system_description': 'Ansul R 102 System',
                    'tank_quantity': '2',
                    'manual_release': '1no station',
                    'base_price': format_price(fs_base_total),
                    'install_price': format_price(fs_install_price),
                    'total_price': format_price(sum(
                        safe_convert_to_float(fs_canopy.get('fire_suppression_data', {}).get('base_price', 0)) +
                        safe_convert_to_float(fs_canopy.get('fire_suppression_data', {}).get('install_price', 0))
                        for fs_canopy in fire_suppression_canopies
                    )),
                    'n9_total': format_price(fs_total)  # Use fs_total instead of n9_total
                } if has_fire_suppression else None,
                'fire_suppression_base_total': format_price(fs_base_total) if has_fire_suppression else format_price(0),
                'fire_suppression_install': format_price(fs_install_price) if has_fire_suppression else format_price(0),
                'has_cladding': bool(area_cladding_items),
                'cladding_items': area_cladding_items,
                'canopy_total': format_price(canopy_total),
                'delivery_total': format_price(delivery_total),
                'commissioning_total': format_price(commissioning_total),
                'cladding_total': format_price(sum(safe_convert_to_float(item['price']) for item in area_cladding_items)),
                'uv_total': format_price(project_data['uv_control_data'].get(display_name, {}).get('total_price', 0)) if display_name in project_data['uv_control_data'] else format_price(0),
                'recoair_total': format_price(project_data['recoair_control_data'].get(display_name, {}).get('total_price', 0)) if display_name in project_data['recoair_control_data'] else format_price(0),
                'area_total': format_price(area_total),
                'total_price': format_price(sheet_total),
                'mua_calculations': {
                    'total_extract_volume': round(total_extract, 3),
                    'required_mua': round(required_mua, 3),
                    'total_mua_volume': round(total_mua, 3),
                    'mua_shortfall': round(mua_shortfall, 3)
                },
                'important_note': important_note
            })
        
        # 
        # 
        # st.write(areas)
        # Calculate totals from all areas
        # Calculate job total excluding RECOAIR prices
        job_total = sum(safe_convert_to_float(area['total_price']) for area in areas)
        # st.write(job_total)
        k9_total = 0.0
        commissioning_total = 0.0  # Add commissioning total tracking
        
        for sheet in project_data['sheets']:
            # Only add to total if sheet has canopies (is used)
            if any(canopy['reference_number'] != 'ITEM' and 
                   canopy['model'] != 'CANOPY TYPE' and
                   canopy['reference_number'] != 'DELIVERY & INSTALLATION'
                   for canopy in sheet['canopies']):
                try:
                    # Get K9 and commissioning totals from sheets
                    sheet_k9 = float(str(sheet['k9_total']).replace(',', ''))
                    sheet_commissioning = float(str(sheet['commissioning_price']).replace(',', ''))
                    
                    k9_total += sheet_k9
                    commissioning_total += sheet_commissioning
                except (ValueError, TypeError):
                    st.write(f"Warning: Could not convert total price for sheet {sheet.get('sheet_name', 'Unknown')}")
                    continue
        
        # Add fire suppression totals to job total
        try:
            fs_n9_total = float(str(project_data['global_fs_n9_total'] or 0).replace(',', ''))
            
        except (ValueError, TypeError):
            st.write("Warning: Could not convert fire suppression N9 total")
        
        # Format totals with commas and 2 decimal places
        job_total_formatted = f"{math.ceil(job_total):,.2f}"
        k9_total_formatted = f"{math.ceil(k9_total):,.2f}"
        commissioning_total_formatted = f"{math.ceil(commissioning_total):,.2f}"
        
        # Add to context
        context = {
            'date': data['Date'],
            'project_number': halton_ref,  # Updated reference format
            'quote_title': quote_title,    # New quote title
            'sales_contact_name': sales_contact_name,
            'contact_number': contact_number,
            'customer': data['Customer'],
            'customer_first_name': customer_first_name,
            'company': data.get('Company', ''),
            'address': address,  # Use the address variable we populated earlier
            'Address': address,  # Also add with uppercase A for template compatibility
            'estimator_name': estimator_name,  # Use extracted name
            'estimator_role': estimator_role,  # Use extracted role
            'scope_items': [],  # Add scope items to context
            'project_data': project_data,  # Add project_data to context first
            'project_name': project_data['sheets'][0]['project_info']['project_name'],
            'location': project_data['sheets'][0]['project_info']['location'],
            'areas': areas,
            'has_water_wash': bool(project_data.get('has_water_wash', False)),
            'has_uv': project_data.get('has_uv_canopy', False) or has_any_uv_canopy,  # Use both global flags for compatibility
            'has_recoair': project_data.get('has_recoair', False),  # Add global flag for RECOAIR
            'has_sdu': has_sdu,  # Add global flag for SDU
            'has_cladding': bool(cladding_items),
            'cladding_items': cladding_items,
            'cladding_total': sum(safe_convert_to_float(item['price']) for item in cladding_items),
            'f_canopy_mua_vols': f_canopy_mua_vols,
            'ww_canopies': ww_canopies,
            'job_total': job_total_formatted,
            'k9_total': k9_total_formatted,
            'commissioning_total': commissioning_total_formatted,  # Add to context with proper formatting
        }
        # st.write(context)
        
        # Debug output to verify water wash canopies in context
        # st.write(f"DEBUG: Water wash canopies in context: {len(ww_canopies)}")
        
        
        # Debug output to verify has_water_wash status
        
        # st.write(context)
       # st.write(context)
        print(context.get('estimator_name'))
        # Collect and organize canopy data for scope of works
        canopy_counts = {}
        for sheet in project_data['sheets']:
            for canopy in sheet['canopies']:
                # Skip invalid or placeholder canopies
                if (canopy['reference_number'] == 'ITEM' or 
                    canopy['model'] == 'CANOPY TYPE' or 
                    not canopy['model'] or 
                    not canopy['configuration'] or 
                    canopy['model'] == "Select..." or 
                    canopy['configuration'] == "Select..."):
                    continue
                    
                key = f"{canopy['model']} {canopy['configuration']}"
                canopy_counts[key] = canopy_counts.get(key, 0) + 1
        
        # Format scope of works
        scope_items = []
        for canopy_type, count in canopy_counts.items():
            scope_items.append(get_canopy_description(canopy_type, count))
        
        # Add any wall cladding from valid canopies only
        wall_cladding_count = sum(1 for sheet in project_data['sheets'] 
                                 for canopy in sheet['canopies'] 
                                 if (canopy['wall_cladding']['type'] and 
                                     canopy['wall_cladding']['type'] != "Select..." and 
                                     canopy['wall_cladding']['type'] == "2M² (HFL)" and
                                     canopy['reference_number'] != 'ITEM' and
                                     canopy['model'] != 'CANOPY TYPE'))
        if wall_cladding_count > 0:
            scope_items.append(f"{wall_cladding_count}no Areas with Stainless Steel Cladding")
        
        # Add scope items to context
        context['scope_items'] = scope_items
        
        # Inside write_to_word_doc function, before processing areas
        # Create a map of fire suppression data by item number
        fs_data = []
        for sheet in project_data['sheets']:
            for canopy in sheet['canopies']:
                if canopy.get('has_fire_suppression') and canopy.get('fire_suppression_data'):
                    fs_data.append({
                        'item_number': canopy['reference_number'],
                        'model': canopy['model'],
                        'system_description': canopy['fire_suppression_data']['system_description'],
                        'tank_quantity': canopy['fire_suppression_data']['tank_quantity'],
                        'manual_release': canopy['fire_suppression_data']['manual_release']
                    })

        # Add fire suppression data to context
        context.update({
            'has_fire_suppression': bool(fs_data),
            'fire_suppression_data': fs_data,
            'fire_suppression_description': """The restaurant fire suppression system is a pre-engineered, wet chemical, cartridge-operated, regulated pressure type with a fixed nozzle agent distribution network. The system is capable of automatic detection and actuation and / or remote manual actuation to provide fire protection to the cooking appliances, canopy exhaust duct and canopy filter plenum. Fire alarm / BMS connections to the release module to be carried out by others."""
        })
        
        # Render the template with the context
        doc.render(context)
        
        # Save the document
        doc.save(output_path)
        return output_path
    except Exception as e:
        st.error(f"Error generating Word document: {str(e)}")
        return None

def generate_word_document(project_data):
    """Generate Word document from project data"""
    try:
        # Check if we have sheets
        if not project_data or 'sheets' not in project_data or not project_data['sheets']:
            st.error("No sheet data found to generate Word document.")
            return None
        
        # Get data from first sheet with safe access
        first_sheet = project_data['sheets'][0]
        project_info = first_sheet.get('project_info', {})
        
        word_data = {
            'Date': project_info.get('date', ''),
            'Project Number': project_info.get('project_number', ''),
            'Sales Contact': project_info.get('sales_estimator', '').split('/')[0] if project_info.get('sales_estimator') else '',
            'Estimator': project_info.get('sales_estimator', ''),
            'Estimator_Name': project_info.get('estimator_name', ''),
            'Estimator_Role': project_info.get('estimator_role', ''),
            'Customer': project_info.get('customer', ''),
            'Company': project_info.get('company', ''),
            'Address': project_info.get('address', '')  # Get address from project info
        }
        #st.write(word_data)
        
        # Generate Word document
        return write_to_word_doc(word_data, project_data)
    except Exception as e:
        st.error(f"Error generating Word document: {str(e)}")
        return None

# Modify save_to_excel function to remove Word document generation
def save_to_excel(data):
    try:
        # Define paths using os.path for cross-platform compatibility
        template_path = os.path.join("resources", "Halton Cost Sheet Jan 2025.xlsx")
        output_path = "output.xlsx"
        
       # st.write("🔍 Starting Excel file generation...")
        
        # Check if template exists
        if not os.path.exists(template_path):
            st.error(f"Template file not found. Please ensure '{template_path}' exists in the resources folder.")
            return
            
        # Load workbook
        workbook = openpyxl.load_workbook(template_path, read_only=False, data_only=False)
        
        # Hide the Lists sheet
        if 'Lists' in workbook.sheetnames:
            workbook['Lists'].sheet_state = 'hidden'
        
        # Write address to LISTS sheet
        write_address_to_lists(workbook, data.get('Address', ''))
        
        # Get all sheets once
        all_sheets = workbook.sheetnames
        canopy_sheets = [sheet for sheet in all_sheets if 'CANOPY' in sheet]
        fire_supp_sheets = [sheet for sheet in all_sheets if 'FIRE SUPP' in sheet]
        edge_box_sheets = [sheet for sheet in all_sheets if 'EBOX' in sheet or 'EDGE BOX' in sheet]
        recoair_sheets = [sheet for sheet in all_sheets if 'RECOAIR' in sheet]  # Add dedicated RECOAIR sheets
        sdu_sheets = [sheet for sheet in all_sheets if 'SDU' in sheet]  # Add dedicated SDU sheets
        
        #st.write(f"📑 Found template sheets: {len(canopy_sheets)} CANOPY, {len(fire_supp_sheets)} FIRE SUPP, {len(edge_box_sheets)} EBOX")
        
        sheet_count = 0
        fs_sheet_count = 0
        ebox_sheet_count = 0
        recoair_sheet_count = 0  # Add counter for RECOAIR sheets
        sdu_sheet_count = 0  # Add counter for SDU sheets
        
        # Create a template cache for fire suppression sheets
        fs_template = None
        
        # Define a dictionary of colors to assign to different floors
        floor_colors = {
            # RGB colors (in hex format)
            # 20 distinct, professional colors without grays
            0: "FF4F81BD",  # Blue
            1: "FF8064A2",  # Purple
            2: "FF8EB33B",  # Green
            3: "FFC0504D",  # Red
            4: "FFFF9900",  # Orange
            5: "FF00B0F0",  # Light Blue
            6: "FF76933C",  # Dark Green
            7: "FFBF8F00",  # Dark Yellow
            8: "FF1F497D",  # Navy Blue
            9: "FF953735",  # Dark Red
            10: "FF4BACC6", # Teal
            11: "FF92D050", # Bright Green
            12: "FFAC639D", # Pink
            13: "FF9966FF", # Violet
            14: "FF538DD5", # Sky Blue
            15: "FFC3D69B", # Light Green
            16: "FFFF6600", # Bright Orange
            17: "FF0070C0", # Strong Blue
            18: "FFC00000", # Dark Red
            19: "FF00B050"  # Emerald Green
        }
        
        # Track which area is which color
        area_color_map = {}
        area_index = 0
        
        # Process each level and its areas
        for level in data['Levels']:  # Changed 'levels' to 'Levels'
            level_name = level['level_name']
            
            for area in level['areas']:
                area_name = area['area_name']
                
                # Create a unique key for each area
                area_key = f"{level_name} - {area_name}"
                
                # Assign a color to this area if not already assigned
                if area_key not in area_color_map:
                    area_color_map[area_key] = floor_colors[area_index % len(floor_colors)]
                    area_index += 1
                    # Debug message to verify areas are getting different colors
                    print(f"Assigned color #{area_index-1} to area: {area_key}")
                
                # Get the color for this area
                tab_color = area_color_map[area_key]
                
                # Check for canopies with fire suppression (ensure 'canopies' key exists and is not empty)
                area_canopies = area.get('canopies', [])
                has_fire_suppression = any(canopy.get('fire_suppression', False) for canopy in area_canopies)
                
                # Check if area has UV-C Control Schedule
                needs_edge_box = area.get('include_uvc', False)
                
                # Check if area includes RECOAIR system
                needs_recoair = area.get('include_recoair', False)
                
                # Check if area includes SDU
                needs_sdu = area.get('include_sdu', False)

                current_canopy_sheet = None # Initialize
                fs_sheet = None # Initialize

                if area_canopies: # Only process canopy sheet if canopies exist for this area
                    if canopy_sheets:
                        sheet_name = canopy_sheets.pop(0)
                        current_canopy_sheet = workbook[sheet_name]
                        sheet_title_display = f"{level_name} - {area_name}" # For B1 cell
                        current_canopy_sheet['B1'] = sheet_title_display
                        
                        # Use sheet_count for CANOPY sheet numbering
                        canopy_sheet_tab_name = f"CANOPY - {level_name} ({sheet_count + 1})"
                        current_canopy_sheet.title = canopy_sheet_tab_name
                        current_canopy_sheet.sheet_state = 'visible'
                        current_canopy_sheet.sheet_properties.tabColor = tab_color
                        
                        # Create fire suppression sheet if needed (associated with these canopies)
                        if has_fire_suppression:
                            if fire_supp_sheets:
                                fs_sheet_name_template = fire_supp_sheets.pop(0)
                                fs_sheet = workbook[fs_sheet_name_template]
                                # FS sheet naming and numbering tied to its corresponding CANOPY sheet
                                new_fs_name = f"FIRE SUPP - {level_name} ({sheet_count + 1})" 
                                fs_sheet.title = new_fs_name
                                fs_sheet.sheet_state = 'visible'
                                fs_sheet.sheet_properties.tabColor = tab_color
                                # fs_sheet_count incremented when an fs_sheet is actually used.
                                # However, fs_sheet_count is not used for naming here to keep it linked with canopy sheet.
                                # If independent numbering for FS sheets is ever needed, fs_sheet_count would be used in title.
                            else:
                                st.warning(f"Not enough FIRE SUPP sheets in template for area {area_name} with canopies needing fire suppression. Skipping FIRE SUPP sheet for this area.")
                        
                        write_to_sheet(current_canopy_sheet, data, level_name, area_name, area_canopies, fs_sheet)
                        add_dropdowns_to_sheet(workbook, current_canopy_sheet, 12)
                        if fs_sheet:
                            add_fire_suppression_dropdown(fs_sheet)
                            fs_sheet_count += 1 # Increment global FS sheet counter if one was actually used.
                        
                        sheet_count += 1 # Increment global CANOPY sheet counter
                    else:
                        st.error(f"Not enough CANOPY sheets in template for area {area_name} which has canopies. Skipping CANOPY sheet for this area.")
                        # Continue to next area to allow processing of other areas.
                        # EBOX/RECOAIR/SDU for this area might still be processed below if needed.
                
                # Handle EBOX sheet if needed (can exist without canopies)
                if needs_edge_box:
                    if edge_box_sheets:
                        ebox_sheet_name_template = edge_box_sheets.pop(0)
                        ebox_sheet = workbook[ebox_sheet_name_template]
                        # Use ebox_sheet_count for EBOX sheet numbering
                        new_ebox_name = f"EBOX - {level_name} ({ebox_sheet_count + 1})"
                        ebox_sheet.title = new_ebox_name
                        ebox_sheet.sheet_state = 'visible'
                        ebox_sheet.sheet_properties.tabColor = tab_color
                        
                        # write_to_sheet for EBOX handles its own title in C1
                        write_to_sheet(ebox_sheet, data, level_name, area_name, [], None, True) # Pass empty canopy list
                        ebox_sheet_count += 1
                    else:
                        st.warning(f"No more EDGE BOX (EBOX) sheets available for UV-C in {area_name}. Skipping EBOX sheet.")
                
                # Handle RECOAIR sheet if needed (can exist without canopies)
                if needs_recoair:
                    if recoair_sheets:
                        recoair_sheet_name_template = recoair_sheets.pop(0)
                        recoair_sheet = workbook[recoair_sheet_name_template]
                        # Use recoair_sheet_count for RECOAIR sheet numbering
                        new_recoair_name = f"RECOAIR - {level_name} ({recoair_sheet_count + 1})"
                        recoair_sheet.title = new_recoair_name
                        recoair_sheet.sheet_state = 'visible'
                        recoair_sheet.sheet_properties.tabColor = tab_color
                        
                        try:
                            recoair_sheet.cell(row=1, column=2, value="RECOAIR CONTROL SYSTEM") # B1
                        except Exception as e:
                            st.warning(f"Could not set title for RECOAIR sheet {new_recoair_name}: {str(e)}")
                        
                        # write_to_sheet for RECOAIR handles its own title in C1
                        write_to_sheet(recoair_sheet, data, level_name, area_name, [], None, True) # Pass empty list
                        
                        item_number = f"1.{recoair_sheet_count + 1:02d}" 
                        recoair_sheet.cell(row=12, column=3, value=item_number)
                        recoair_sheet['D13'] = "Model" 
                        add_recoair_internal_external_dropdowns(recoair_sheet)
                        recoair_sheet_count += 1
                    else:
                        st.warning(f"No more RECOAIR sheets available for RECOAIR System in {area_name}. Skipping RECOAIR sheet.")
                
                # Handle SDU sheet if needed (can exist without canopies)
                if needs_sdu:
                    if sdu_sheets:
                        sdu_sheet_name_template = sdu_sheets.pop(0)
                        sdu_sheet = workbook[sdu_sheet_name_template]
                        # Use sdu_sheet_count for SDU sheet numbering
                        new_sdu_name = f"SDU - {level_name} ({sdu_sheet_count + 1})"
                        sdu_sheet.title = new_sdu_name
                        sdu_sheet.sheet_state = 'visible'
                        sdu_sheet.sheet_properties.tabColor = tab_color
                        
                        try:
                            # SDU sheet's B1 title should be the area name
                            sdu_sheet.cell(row=1, column=2, value=f"{level_name} - {area_name}") # B1
                        except Exception as e:
                            st.warning(f"Could not set title for SDU sheet {new_sdu_name}: {str(e)}")

                        # write_to_sheet for SDU handles its own title in C1
                        write_to_sheet(sdu_sheet, data, level_name, area_name, [], None, True) # Pass empty list
                        sdu_sheet_count += 1
                    else:
                        st.warning(f"No more SDU sheets available for SDU in {area_name}. Skipping SDU sheet.")
                
        # Make sure the JOB TOTAL sheet has no tab color to differentiate it
        if 'JOB TOTAL' in workbook.sheetnames:
            # We're not setting a tab color, which will keep it as default
            # Ensure it's visible
            workbook['JOB TOTAL'].sheet_state = 'visible'
        
        # Reorganize the sheets so related sheets are grouped by area
        organize_sheets_by_area(workbook)
        
        # Hide the Lists sheet at the end of the process to ensure it stays hidden
        if 'Lists' in workbook.sheetnames:
            workbook['Lists'].sheet_state = 'hidden'
        
        # Save the workbook
        workbook.save(output_path)
        st.success(f"""✅ Successfully created:
- {sheet_count} CANOPY sheets
- {fs_sheet_count} FIRE SUPPRESSION sheets
- {ebox_sheet_count} EDGE BOX sheets
- {recoair_sheet_count} RECOAIR sheets
- {sdu_sheet_count} SDU sheets

All sheets are color-coded by area for easy identification.""")
        
        # Provide download button
        with open(output_path, "rb") as file:
            st.download_button(
                label="📥 Download Excel file",
                data=file,
                file_name="project_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
    except FileNotFoundError:
        st.error("❌ Template file not found in resources folder!")
    except Exception as e:
        st.error(f"❌ An error occurred: {str(e)}")
        raise e

def extract_sheet_data(sheet):
    """Extract data from a single sheet"""
    # Skip F24 sheets
    if 'F24' in sheet.title:
        return None
    
    # Also check B1 value for F24
    b1_value = sheet['B1'].value
    if b1_value and 'F24' in str(b1_value):
        return None
    
    # Get estimator info from Lists sheet
    list_sheet = sheet.parent['Lists']
    estimator_name = list_sheet['Z1'].value
    estimator_role = list_sheet['Z2'].value
    
    # Get address from LISTS sheet
    address = read_address_from_lists(sheet.parent)
    
    # Get sheet name from B1 (Floor Name - Area)
    display_name = sheet['B1'].value or sheet.title
    
    # Parse customer and company from C5
    customer_value = sheet['C5'].value or ''
    if '(' in customer_value and ')' in customer_value:
        # Split into customer and company if in format "Customer (Company)"
        customer = customer_value.split('(')[0].strip()
        company = customer_value.split('(')[1].rstrip(')').strip()
    else:
        customer = customer_value
        company = ''
    
    # Get delivery and installation price from P182
    delivery_install_price = safe_convert_to_float(sheet['P182'].value)
    
    # Get K9 total from sheet
    k9_total = math.ceil(safe_convert_to_float(sheet['K9'].value))
    
    # Get commissioning price from N193
    commissioning_price = safe_convert_to_float(sheet['N193'].value)
    
    data = {
        'sheet_name': display_name,  # Use B1 value instead of sheet title
        'revision': sheet['O7'].value,
        'delivery_install': delivery_install_price,
        'total_price': math.ceil(safe_convert_to_float(sheet['N9'].value)),
        'k9_total': k9_total,
        'commissioning_price': commissioning_price,  # Store raw value
        'project_info': {
            'project_number': sheet['C3'].value,
            'customer': customer,
            'company': company,
            'sales_estimator': sheet['C7'].value,
            'project_name': sheet['G3'].value,
            'location': sheet['G5'].value,
            'date': sheet['G7'].value,
            'estimator_name': estimator_name,
            'estimator_role': estimator_role,
            'address': address  # Add the address from Lists sheet to project_info
        },
        'canopies': [],
        'fire_suppression_items': []  # Add new list for fire suppression items
    }
    
    # Extract canopy data - only process up to row 181 for canopies
    row = 12  # Starting row for first canopy
    while row <= 181:  # Processing canopies
        reference_number = sheet[f'B{row}'].value
        
        # Skip this canopy if reference number contains "ITEM" or is empty
        if not reference_number or 'ITEM' in str(reference_number).upper().strip():
            row += 17
            continue
            
        model = sheet[f'D{row + 2}'].value
        configuration = sheet[f'C{row + 2}'].value
        
        # Only process if we have valid model and configuration
        if (model and model != "Select..." and 
            configuration and configuration != "Select..."):
            
            dim_row = row + 2
            pa_row = row + 10  # F22 for first canopy
            cladding_row = row + 7
            water_row = row + 13  # F25 for first canopy water wash values
            
            # Get K9 value for this canopy
            canopy_k9 = safe_convert_to_float(sheet[f'K{row}'].value)
            
            # Get fire suppression data
            fs_system = sheet[f'C{row + 4}'].value  # C16 - System description
            fs_tank_qty = sheet[f'C{row + 4}'].value  # C16 - Tank quantity
            
            # Format fire suppression data
            has_fire_suppression = False
            fire_suppression_base_price = 0
            fire_suppression_install = 0
            fire_suppression_n9 = 0
            tank_quantity = "1"  # Default value
            fs_canopy_count = 0  # Initialize fire suppression canopy counter
            per_canopy_delivery = 0  # Initialize per-canopy delivery share
            canopy_total_price = 0  # Initialize total price for canopy
            
            # Look for matching fire suppression sheet
            for fs_sheet_name in sheet.parent.sheetnames:
                if 'FIRE SUPP' in fs_sheet_name and 'F24' not in fs_sheet_name:
                    fs_sheet = sheet.parent[fs_sheet_name]
                    # Check if this is the correct fire suppression sheet for this area
                    if fs_sheet['B1'].value == sheet['B1'].value:
                        # Find matching item in fire suppression sheet
                        fs_row = 12
                        while fs_row <= 181:
                            if fs_sheet[f'B{fs_row}'].value == sheet[f'B{row}'].value:
                                has_fire_suppression = True
                                fire_suppression_base_price = safe_convert_to_float(fs_sheet[f'N{fs_row}'].value)
                                fire_suppression_base_price = math.ceil(fire_suppression_base_price)
                                
                                # Get tank quantity from C16 in fire suppression sheet
                                system_desc = fs_sheet[f'C{fs_row + 4}'].value or ""
                                if system_desc == "FIRE SUPPRESSION":
                                    tank_quantity = "-"
                                else:
                                    # Try to extract number from system description
                                    try:
                                        # Split the string and get the first part which should be the number
                                        first_word = system_desc.split()[0]
                                        if first_word.isdigit():
                                            tank_quantity = first_word
                                        else:
                                            tank_quantity = "1"
                                    except (AttributeError, IndexError):
                                        tank_quantity = "1"
                                
                                # Get N182 value from the fire suppression sheet
                                fire_suppression_install = safe_convert_to_float(fs_sheet['N182'].value)
                                
                                # Count the total number of valid fire suppression canopies (for delivery cost sharing)
                                fs_canopy_count = 0
                                for check_row in range(12, 182, 17):
                                    ref_value = fs_sheet[f'B{check_row}'].value
                                    # Only count rows with valid reference numbers (not empty, not just whitespace, and not "ITEM" or "DELIVERY")
                                    if (ref_value and str(ref_value).strip() and 
                                        "ITEM" not in str(ref_value).upper() and 
                                        "DELIVERY" not in str(ref_value).upper()):
                                        fs_canopy_count += 1
                                
                                # Calculate per-canopy share of delivery
                                per_canopy_delivery = 0
                                if fs_canopy_count > 0:
                                    per_canopy_delivery = math.ceil(fire_suppression_install / fs_canopy_count)
                                
                                # Calculate total price for this canopy (base + per-canopy delivery)
                                canopy_total_price = fire_suppression_base_price + per_canopy_delivery
                                
                                # Get N9 value from the fire suppression sheet
                                fire_suppression_n9 = safe_convert_to_float(fs_sheet['N9'].value)

                                # Debug information
                                print(f"--- Fire Suppression Pricing for {sheet[f'B{row}'].value} ---")
                                print(f"Fire Suppression Sheet: {fs_sheet_name}")
                                print(f"Fire Suppression Row: {fs_row}")
                                print(f"Reference number: {fs_sheet[f'B{fs_row}'].value}")
                                print(f"Base price from N{fs_row}: £{fire_suppression_base_price}")
                                print(f"Delivery cost from N182: £{fire_suppression_install}")
                                print(f"Total FS canopies: {fs_canopy_count}")
                                print(f"Per-canopy delivery share: £{per_canopy_delivery}")
                                print(f"Total price for canopy: £{canopy_total_price}")
                                print(f"Total for sheet (N9): £{fire_suppression_n9}")
                                print(f"Setting has_fire_suppression = True for canopy {sheet[f'B{row}'].value}")
                                
                               #st.write(fire_suppression_n9)
                                break
                            fs_row += 17
                        if has_fire_suppression:
                            break

            fire_suppression_data = {
                'system_description': fs_system or f"Ansul R102 System to cover item {sheet[f'B{row}'].value}",
                'tank_quantity': tank_quantity,
                'manual_release': '1no station',
                'base_price': math.ceil(fire_suppression_base_price),
                'install_price': math.ceil(per_canopy_delivery),  # Store per-canopy delivery share
                'total_price': math.ceil(fire_suppression_base_price + per_canopy_delivery)  # Base + per-canopy delivery
            } if has_fire_suppression else None
            
            canopy_price = safe_convert_to_float(sheet[f'P{row}'].value)  # Base canopy price
            canopy_price = math.ceil(canopy_price)  # Round up to nearest pound
            
            cladding_price = safe_convert_to_float(sheet[f'N{row + 7}'].value)  # Cladding price if exists
            emergency_lighting = sheet[f'P{row + 1}'].value or "2No, @ £100.00 if required"  # Emergency lighting text
            
            # Get water wash values
            cws_2bar = sheet[f'F{row + 13}'].value  # F25 - Cold water supply at 2 bar
            hws_2bar = sheet[f'F{row + 14}'].value  # F26 - Hot water supply at 2 bar
            hws_storage = sheet[f'F{row + 15}'].value  # F27 - Hot water storage
            # st.write(f"Water Wash Values: {cws_2bar}, {hws_2bar}, {hws_storage}")
            # Get water wash prices
            ww_price = safe_convert_to_float(sheet[f'P{row + 13}'].value)  # P25 - Water wash price
            ww_control_price = safe_convert_to_float(sheet[f'P{row + 14}'].value)  # P26 - Control panel price
            ww_install_price = safe_convert_to_float(sheet[f'P{row + 15}'].value)  # P27 - Installation price
            
            wall_cladding = {
                'type': sheet[f'C{cladding_row}'].value,
                'width': sheet[f'Q{cladding_row}'].value or 0,
                'height': sheet[f'R{cladding_row}'].value or 0,
                'positions': (sheet[f'S{cladding_row}'].value or '').split(',') if sheet[f'S{cladding_row}'].value else [],
                'price': math.ceil(safe_convert_to_float(cladding_price))
            }
            
            # Get MUA VOL from Q22 for F-type canopies
            mua_vol = '-'
            model_value = sheet[f'D{row + 2}'].value
            if model_value:
                model_str = str(model_value)  # Convert to string first
                if 'F' in model_str.upper():
                    # Get extract volume
                    ext_vol = sheet[f'I{dim_row}'].value or 0
                    # Get MAX SUPPLY value from H22 + offset
                    max_supply_cell = sheet[f'H{22 + ((row - 12) // 17) * 17}'].value or ''
                   # st.write(max_supply_cell)
                    # Calculate 85% of extract volume and round to 2 decimal places
                    calculated_mua = round(float(ext_vol) * 0.85, 2) if ext_vol != '-' else 0
                    
                    # Parse MAX SUPPLY value (could be a string with (MAX) or just a number)
                    # Check if the cell has a value at all
                    if max_supply_cell:
                        # Convert to string first to handle both string and numeric cell values
                        max_supply_str = str(max_supply_cell)
                        # Remove (MAX) if present
                        max_supply_str = max_supply_str.replace('(MAX)', '').strip()
                        try:
                            max_supply = float(max_supply_str)
                           # st.write(f"MAX: {max_supply}")
                            #st.write(f"85%: {calculated_mua}")
                            # IMPORTANT: If MAX is less than calculated MUA (85% of extract), we MUST use the MAX value
                            # The MAX value represents the system limit that cannot be exceeded
                            if max_supply < calculated_mua:
                                mua_vol = str(round(max_supply, 2)) + '(MAX)'
                            else:
                                mua_vol = str(calculated_mua)
                        except (ValueError, TypeError):
                            # If we can't convert to float, use calculated MUA
                            mua_vol = str(calculated_mua)
                    else:
                        # If no MAX value, use calculated MUA
                        mua_vol = str(calculated_mua)
                    #st.write(mua_vol)
            
            # Inside the canopy processing loop, update the ext_static and lighting handling:
            model = sheet[f'D{row + 2}'].value or '-'
            lighting_value = sheet[f'C{row + 3}'].value or '-'
            print(lighting_value)
            if lighting_value == 'LIGHT SELECTION':
                lighting = '-'
            else:
                # Convert lighting value to simplified format
                if 'LED STRIP' in str(lighting_value).upper():
                    lighting = 'LED Strip'
                elif 'SPOTS' in str(lighting_value).upper():
                    lighting = 'LED Spots'
                else:
                    lighting = str(lighting_value)
                
            # Handle ext_static based on model
            ext_static_value = sheet[f'F{pa_row}'].value or '-'
            if model == 'CXW':
                ext_static = 45  # Always 45 for CXW models
            elif ext_static_value != '-':
                # Remove 'Pa' and convert to string
                ext_static_str = str(ext_static_value).replace('Pa', '').strip()
                try:
                    # Convert to float and round
                    ext_static = round(float(ext_static_str))
                except (ValueError, TypeError):
                    ext_static = '-'
            else:
                ext_static = '-'
            
            # Handle supply_static based on model
            supply_static_value = sheet[f'L{dim_row}'].value
            if 'F' in str(model).upper():
                # Default value for F-type canopies is 45
                supply_static = 45
                # Use value from sheet if available
                if supply_static_value and supply_static_value != '-':
                    try:
                        supply_static = round(float(str(supply_static_value).replace('Pa', '').strip()))
                    except (ValueError, TypeError):
                        # Keep default of 45 if conversion fails
                        supply_static = 45
            else:
                # Non-F canopies get '-' for supply_static
                supply_static = '-'
            
            # Create canopy data dictionary and append to list
            canopy_data = {
                'reference_number': reference_number,
                'model': model,
                'configuration': configuration,
                'length': sheet[f'E{dim_row}'].value or '-',
                'width': sheet[f'F{dim_row}'].value or '-',
                'height': sheet[f'G{dim_row}'].value or '-',
                'sections': sheet[f'H{dim_row}'].value or '-',
                'ext_vol': sheet[f'I{dim_row}'].value or '-',
                'ext_static': ext_static if ext_static != 0 else '-',
                'supply_static': supply_static if supply_static != 0 else '-',
                'base_price': format_price(canopy_price),
                'k9_price': canopy_k9,
                'wall_cladding': wall_cladding,
                'has_fire_suppression': has_fire_suppression,
                'fire_suppression_data': fire_suppression_data,
                'mua_vol': mua_vol,
                'lighting': lighting,
                'special_works_1': sheet[f'C{row + 4}'].value or '-',
                'special_works_2': sheet[f'C{row + 5}'].value or '-',
                'bim_revit': sheet[f'C{row + 6}'].value or '-',
                'emergency_lighting': emergency_lighting or '-',
                'cws_2bar': cws_2bar or '-',
                'hws_2bar': hws_2bar or '-',
                'hws_storage': hws_storage or '-',
                'ww_price': ww_price,
                'ww_control_price': ww_control_price,
                'ww_install_price': ww_install_price,
                'control_panel': sheet[f'C{row + 13}'].value or '-',  # C25
                'ww_pods': sheet[f'C{row + 14}'].value or '-',        # C26
                'ww_control': sheet[f'C{row + 15}'].value or '-',      # C27
                'is_water_wash': 'CMWF' in str(model).upper() or 'CMWI' in str(model).upper()  # Add flag for water wash canopies
            }
            
            data['canopies'].append(canopy_data)
            
            # If this canopy has fire suppression, add it to fire_suppression_items
            if has_fire_suppression:
                data['has_fire_suppression'] = True  # Set the flag at the data level
                # Store the fire suppression delivery cost at sheet level for later calculation
                data['fire_suppression_install'] = fire_suppression_install
                
                data['fire_suppression_items'].append({
                    'reference_number': reference_number,
                    'model': model,
                    'has_fire_suppression': True,  # Explicitly set this flag
                    'fire_suppression_data': fire_suppression_data
                })
                
                # Add a debug print to troubleshoot fire suppression detection
                print(f"DEBUG: Added canopy {reference_number} to fire_suppression_items")
                print(f"DEBUG: fire_suppression_data = {fire_suppression_data}")
                print(f"DEBUG: fire_suppression_install = {fire_suppression_install}")
            
            # Check if this is a water wash canopy and set data flag immediately
            if 'CMWF' in str(model).upper() or 'CMWI' in str(model).upper():
                data['has_water_wash'] = True
                # Add to water wash items list if not already present
                if 'water_wash_items' not in data:
                    data['water_wash_items'] = []
                
                # Store detailed water wash information
                data['water_wash_items'].append({
                    'reference_number': reference_number,
                    'model': model,
                    'cws_2bar': cws_2bar or '-',
                    'hws_2bar': hws_2bar or '-',
                    'hws_storage': hws_storage or '-',
                    'ww_price': ww_price,
                    'ww_control_price': ww_control_price,
                    'ww_install_price': ww_install_price,
                    'control_panel': sheet[f'C{row + 13}'].value or '-',
                    'ww_pods': sheet[f'C{row + 14}'].value or '-',
                    'ww_control': sheet[f'C{row + 15}'].value or '-'
                })
                
                # Debug output for water wash canopies
                # st.write(f"DEBUG: Added water wash canopy: {reference_number} ({model})")
            
        row += 17  # Move to next canopy section
    
    # After processing all canopies, output water wash summary
    
    
    # Calculate sheet K9 total from individual canopy K9 values
    calculated_k9_total = sum(math.ceil(canopy['k9_price']) for canopy in data['canopies'])
    # Use calculated total if sheet total is 0 or missing
    if k9_total == 0:
        k9_total = calculated_k9_total
    data['k9_total'] = k9_total
    
    # Calculate MUA requirements and shortfall
    total_extract_volume = 0
    total_mua_volume = 0
    
    # Sum up all extract volumes and MUA volumes
    for canopy in data['canopies']:
        # Get extract volume
        ext_vol = canopy.get('ext_vol', 0)  # Use get() with default value
        if ext_vol and ext_vol != '-':
            try:
                # Convert to float and add to total
                ext_vol_float = float(str(ext_vol).replace(',', ''))
                total_extract_volume += ext_vol_float
            except (ValueError, TypeError):
                pass
        
        # Get MUA volume
        mua_vol = canopy.get('mua_vol', '-')  # Use get() with default value
        if mua_vol and mua_vol != '-':
            try:
                # Remove (MAX) from the mua_vol string if present
                mua_vol_clean = str(mua_vol).replace('(MAX)', '').strip()
                # Convert to float and add to total
                mua_vol_float = float(mua_vol_clean.replace(',', ''))
                total_mua_volume += mua_vol_float
            except (ValueError, TypeError):
                pass
    
    # Calculate required MUA (85% of total extract)
    required_mua = round(total_extract_volume * 0.85, 3)
    
    # Calculate shortfall (required MUA minus actual MUA volume)
    mua_shortfall = round(required_mua - total_mua_volume, 3)
    
    # Add MUA calculations to data structure
    data['mua_calculations'] = {
        'total_extract_volume': round(total_extract_volume, 3),
        'required_mua': required_mua,
        'total_mua_volume': round(total_mua_volume, 3),
        'mua_shortfall': mua_shortfall
    }
    
    # Format the important note
    data['important_note'] = (
        f"Important Note: - The make-up air flows shown above are the maximum that we can introduce through the "
        f"canopy. This should be equal to approximately 85% of the extract i.e. {required_mua}m³/s\n"
        f"In this instance it only totals {round(total_mua_volume, 3)}m³/s therefore the shortfall of "
        f"{mua_shortfall}m³/s must be introduced through ceiling grilles or diffusers, by others.\n"
        f"If you require further guidance on this, please do not hesitate to contact us."
    )
    
    # Get sheet totals using safe conversion
    data['total_price'] = math.ceil(safe_convert_to_float(sheet['N9'].value))
    data['delivery_install'] = math.ceil(safe_convert_to_float(sheet['P182'].value))
    data['commissioning_price'] = safe_convert_to_float(sheet['N193'].value)
    
    # Extract fire suppression data if this is a FIRE SUPP sheet
    if 'FIRE SUPP' in sheet.title:
        # Process FIRE SUPP sheet
        
        # Get K9 and N9 cell values
        k9_cell_value = safe_convert_to_float(sheet['K9'].value)
        n9_cell_value = safe_convert_to_float(sheet['N9'].value)
        
        # Process fire suppression items
        fire_suppression_items = []
        for row in range(1, sheet.max_row + 1):
            item_number = sheet[f'A{row}'].value
            if item_number and str(item_number) != 'ITEM':
                k9_value = safe_convert_to_float(sheet[f'K{row}'].value)
                n9_value = safe_convert_to_float(sheet[f'N{row}'].value)
                fs_install = n9_cell_value
                
                system_desc = sheet[f'C{row}'].value
                manual_release = sheet[f'D{row}'].value
                tank_qty = sheet[f'E{row}'].value
                
                if system_desc and manual_release and tank_qty:
                    fs_item = {
                        'item_number': str(item_number),
                        'system_description': system_desc,
                        'manual_release': manual_release,
                        'tank_quantity': tank_qty,
                        'fire_suppression_install': fs_install,
                        'k9_value': k9_value,
                        'n9_value': n9_value
                    }
                    fire_suppression_items.append(fs_item)
        
        # Store fire suppression totals in data
        data['fs_k9_total'] = k9_cell_value
        data['fs_n9_total'] = n9_cell_value
        data['fire_suppression_items'] = fire_suppression_items
    
    return data

def safe_convert_to_float(value):
    """Safely convert a value to float, handling various formats and special cases"""
    if not value:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Convert to string and clean up
    str_value = str(value).upper().strip()
    # Check for special values
    if 'SELECT' in str_value or str_value == '-':
        return 0.0
    try:
        # Remove commas and convert to float
        return float(str_value.replace(',', ''))
    except (ValueError, TypeError):
        return 0.0

def read_excel_file(uploaded_file):
    """Read and process uploaded Excel file"""
    workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
    
    # Get all sheets, excluding F24 sheets
    canopy_sheets = [sheet for sheet in workbook.sheetnames if 'CANOPY' in sheet and 'F24' not in sheet]
    fire_supp_sheets = [sheet for sheet in workbook.sheetnames if 'FIRE SUPP' in sheet and 'F24' not in sheet]
    edge_box_sheets = [sheet for sheet in workbook.sheetnames if ('EBOX' in sheet or 'EDGE BOX' in sheet) and 'F24' not in sheet]
    recoair_sheets = [sheet for sheet in workbook.sheetnames if 'RECOAIR' in sheet and 'F24' not in sheet]
    sdu_sheets = [sheet for sheet in workbook.sheetnames if 'SDU' in sheet and 'F24' not in sheet]
    
    # Initialize project data structure
    project_data = {
        'sheets': [],
        'uv_control_data': {},
        'recoair_control_data': {},
        'has_uv_canopy': False,
        'has_recoair': False,
        'has_water_wash': False,
        'has_sdu': False,  # New flag for SDU
        'water_wash_canopies': [],  # Initialize water wash canopies list
        'has_fire_suppression': False,
        'wall_cladding': {},
        'global_k9_total': 0,
        'global_commissioning_total': 0,
        'global_n9_total': 0,
        'global_fs_k9_total': 0,
        'global_fs_n9_total': 0,
        'global_uv_k9_total': 0,
        'global_uv_n9_total': 0,
        'global_recoair_k9_total': 0,
        'global_recoair_n9_total': 0,
        'global_recoair_delivery_total': 0,    # New global for RECOAIR delivery total
        'global_recoair_commissioning_total': 0,  # New global for RECOAIR commissioning total
        'recoair_flat_pack_details': []  # New list to store flat pack details for each RECOAIR sheet
    }
    
    # Check for area-specific SDU sheets
    for sheet_name in sdu_sheets:
        sheet = workbook[sheet_name]
        b1_value = sheet['B1'].value
        # If B1 doesn't contain "SERVICE DISTRIBUTION UNIT COST", it's an area-specific SDU
        if b1_value and 'SERVICE DISTRIBUTION UNIT COST' not in str(b1_value):
            project_data['has_sdu'] = True
            break
            
    # Process each sheet
    for sheet_name in canopy_sheets + fire_supp_sheets:
        sheet = workbook[sheet_name]
        sheet_data = extract_sheet_data(sheet)
        
        # Skip if sheet_data is None (F24 sheet)
        if sheet_data is None:
            continue
            
        # Check if this sheet has water wash canopies
        if sheet_data.get('has_water_wash', False):
            project_data['has_water_wash'] = True
            
            # Find and collect water wash canopies
            for canopy in sheet_data['canopies']:
                if canopy.get('is_water_wash', False) or 'CMWF' in str(canopy.get('model', '')).upper() or 'CMWI' in str(canopy.get('model', '')).upper():
                    project_data['water_wash_canopies'].append(canopy)
                    # st.write(f"DEBUG: Added water wash canopy to project_data: {canopy.get('reference_number')} ({canopy.get('model')})")
                    # st.write(f"DEBUG: Water wash data: CWS={canopy.get('cws_2bar')}, HWS={canopy.get('hws_2bar')}, Storage={canopy.get('hws_storage')}")
            
            # Debug output for water wash detection
            # st.write(f"DEBUG: Found water wash canopies in sheet {sheet_data['sheet_name']}: {len([c for c in sheet_data['canopies'] if c.get('is_water_wash', False) or 'CMWF' in str(c.get('model', '')).upper() or 'CMWI' in str(c.get('model', '')).upper()])}")
            
        # If this is a canopy sheet, extract cladding data by area and check for UV canopies
        if 'CANOPY' in sheet_name:
            display_name = sheet_data['sheet_name']
            
            # Check for UV canopies in this sheet
            for canopy in sheet_data['canopies']:
                # Skip invalid canopies
                if (canopy.get('reference_number') == 'ITEM' or 
                    canopy.get('model') == 'CANOPY TYPE' or
                    not canopy.get('model') or
                    canopy.get('model') == "Select..." or
                    canopy.get('reference_number') == 'DELIVERY & INSTALLATION'):
                    continue
                
                # Check if any canopy has UV in model or configuration
                model = str(canopy.get('model', '')).upper() if canopy.get('model') is not None else ''
                config = str(canopy.get('configuration', '')).upper() if canopy.get('configuration') is not None else ''
                
                if 'UV' in model or 'UV' in config:
                    project_data['has_uv_canopy'] = True
            
            # Collect cladding items for this area
            area_cladding_items = []
            for canopy in sheet_data['canopies']:
                if (canopy.get('wall_cladding', {}).get('type') and 
                    canopy['wall_cladding']['type'] != "Select..." and
                    canopy['wall_cladding'].get('positions')):
                    # Add cladding item
                    area_cladding_items.append({
                        'item_no': canopy['reference_number'],
                        'description': f"Cladding below Item {canopy['reference_number']}, supplied and installed",
                        'width': canopy['wall_cladding']['width'],
                        'height': canopy['wall_cladding']['height'],
                        'positions': canopy['wall_cladding']['positions'],
                        'price': safe_convert_to_float(canopy['wall_cladding'].get('price', 0))  # Store as numeric value, not formatted string
                    })
            
            # Store cladding items for this area if any exist
            if area_cladding_items:
                project_data['wall_cladding'][display_name] = area_cladding_items
        
        # If this is a fire suppression sheet, check if it has valid items before adding to totals
        if 'FIRE SUPP' in sheet_name:
            # Check if the sheet has any valid fire suppression items
            has_valid_items = any(
                item.get('item_number') and 
                str(item['item_number']).strip().upper() != 'ITEM' and
                str(item['item_number']).strip() != 'DELIVERY & INSTALLATION'
                for item in sheet_data['fire_suppression_items']
            )
            
            if has_valid_items:
                # Add K9 and N9 totals from the sheet data
                project_data['global_fs_k9_total'] += sheet_data.get('fs_k9_total', 0)
                project_data['global_fs_n9_total'] += sheet_data.get('fs_n9_total', 0)
            else:
                continue
        
        # Only add sheet if it has canopies or is a used fire suppression sheet
        has_canopies = any(
            canopy.get('reference_number') and 
            canopy['reference_number'] != 'ITEM' and 
            canopy.get('model') and 
            canopy['model'] != 'CANOPY TYPE' and
            canopy['reference_number'] != 'DELIVERY & INSTALLATION'
            for canopy in sheet_data['canopies']
        )
        
        if has_canopies or ('FIRE SUPP' in sheet_name and sheet_data['fire_suppression_items']):
            project_data['sheets'].append(sheet_data)
    
    # Process EBOX sheets to get UV-C Control Schedule prices
    #st.write(edge_box_sheets)
    for sheet_name in edge_box_sheets:
        sheet = workbook[sheet_name]
        
        # Skip if this is an F24 sheet
        if 'F24' in sheet_name:
            continue
            
        # Also check B1 value for F24
        b1_value = sheet['B1'].value
        if b1_value and 'F24' in str(b1_value):
            continue
            
        area_name = sheet['C1'].value or sheet.title  # Get area name from C1 for EBOX sheets
        # st.write(area_name)
        if area_name:
            # Skip if area name contains F24
            if 'F24' in str(area_name):
                continue
                
            # Get N9 and K9 prices from EBOX sheet
            n9_price = safe_convert_to_float(sheet['N9'].value)
            k9_price = safe_convert_to_float(sheet['K9'].value)
            total_price = safe_convert_to_float(sheet['N9'].value)  # Total price is same as N9 for EBOX sheets
            
            # Add to global totals only if n9_price > 0 (valid data exists)
            if n9_price > 0:
                project_data['global_uv_k9_total'] += k9_price
                project_data['global_uv_n9_total'] += n9_price
                
                project_data['uv_control_data'][area_name] = {
                    'k9_price': k9_price,
                    'n9_price': n9_price,
                    'total_price': total_price  # Add total price to UV control data
                }
                project_data['has_uv_canopy'] = True
    
    # Process RECOAIR sheets to get RECOAIR system prices
    #st.write(recoair_sheets)
    
    # Reset global RECOAIR totals to avoid accumulation
    project_data['global_recoair_k9_total'] = 0
    project_data['global_recoair_n9_total'] = 0
    
    # Track processed areas to avoid duplicates
    processed_recoair_areas = set()
    
    # Initialize RECOAIR models list and totals
    project_data['recoair_models'] = []
    project_data['recoair_subtotal'] = 0
    project_data['recoair_total_price'] = 0
    
    for sheet_name in recoair_sheets:
        sheet = workbook[sheet_name]
        
        # Skip if this is an F24 sheet
        if 'F24' in sheet_name:
            continue
            
        # Also check B1 value for F24 or if it's not a RECOAIR sheet
        b1_value = sheet['B1'].value
        if b1_value and ('F24' in str(b1_value) or 'RECOAIR' not in str(b1_value).upper()):
            continue
            
        area_name = sheet['C1'].value or sheet.title  # Get area name from C1 for RECOAIR sheets
       # st.write(area_name)
        if area_name:
            # Skip if area name contains F24
            if 'F24' in str(area_name):
                # st.write(f"Skipping {sheet_name} (Area name contains F24: {area_name})")
                continue
                
            # Skip if we've already processed this area
            if area_name in processed_recoair_areas:
                # st.write(f"Skipping duplicate RECOAIR area: {area_name} in sheet {sheet_name}")
                continue
            
            commissioning_price = safe_convert_to_float(sheet['N46'].value) if sheet['N46'].value else 0
            
            # Get delivery price from P36
            delivery_price = safe_convert_to_float(sheet['P36'].value) if sheet['P36'].value else 0
            
            # Get K9 and N9 prices
            k9_price = safe_convert_to_float(sheet['K9'].value) if sheet['K9'].value else 0
            n9_price = safe_convert_to_float(sheet['N9'].value) if sheet['N9'].value else 0
            total_price = n9_price  # Total price is same as N9 for RECOAIR sheets
            
            # Get flat pack price from cell N40
            flat_pack_price = safe_convert_to_float(sheet['N40'].value) if sheet['N40'].value else 0
            if flat_pack_price > 0:
                # Round the flat pack price to ceiling
                flat_pack_price = math.ceil(flat_pack_price)
                
                # Add to flat pack details with the area name and item number
                sheet_item_number = sheet['C12'].value
                if not sheet_item_number:
                    sheet_item_number = "1.01"  # Default if missing
                
                # Store flat pack details
                flat_pack_item = {
                    'area_name': area_name,
                    'item_number': sheet_item_number,
                    'flat_pack_price': flat_pack_price
                }
                project_data['recoair_flat_pack_details'].append(flat_pack_item)
            
            # Add to global totals
            project_data['global_recoair_delivery_total'] += delivery_price
            project_data['global_recoair_commissioning_total'] += commissioning_price
            project_data['global_recoair_k9_total'] += k9_price
            project_data['global_recoair_n9_total'] += n9_price
            
            # Add to global totals
            # project_data['global_recoair_k9_total'] += k9_price
            # project_data['global_recoair_n9_total'] += n9_price
            
            # Get the item number from C12 for the entire RECOAIR unit
            sheet_item_number = sheet['C12'].value
            if not sheet_item_number:
                sheet_item_number = "1.01"  # Default if missing
            
            # Process RECOAIR models from rows 14-28
            # st.write(f"DEBUG - Processing RECOAIR models for area: {area_name}")
            recoair_models = []
            
            for row in range(14, 29):
                qty = sheet[f'E{row}'].value
                if qty and safe_convert_to_float(qty) >= 1:
                    # Use the item number from C12 for all models in this sheet
                    item_number = sheet_item_number
                    
                    # Model is in column C (same as before)
                    model = sheet[f'C{row}'].value
                    orientation_volume = sheet[f'D{row}'].value  # This contains text like "VERTICAL 2.25M3/s"
                    width = sheet[f'F{row}'].value
                    length = sheet[f'G{row}'].value
                    height = sheet[f'H{row}'].value
                    price = sheet[f'N{row}'].value
                    location = sheet[f'I{row}'].value  # Get location from column I

                    # Use a default location if none is provided
                    if not location or location == "-" or str(location).strip() == "":
                        location = "Internal"  # Default location
                    
                    # Format the price to have 2 decimal places, rounded up
                    if price:
                        try:
                            # Convert to float, round up to nearest whole number, format with 2 decimal places
                            price_float = safe_convert_to_float(price)
                            price_rounded = math.ceil(price_float)
                            
                            # Format with commas for thousands and 2 decimal places
                            price_formatted = f"{price_rounded:,.2f}"
                        except (ValueError, TypeError):
                            price_formatted = price
                    else:
                        price_formatted = "0.00"
                    
                    # Process the model name
                    if model:
                        # Store original model name
                        original_model = model
                        
                        # Convert model name to string if it's not already
                        model = str(model)
                        
                        # Convert RA prefix to RAH if needed
                        if model.upper().startswith('RA') and not model.upper().startswith('RAH'):
                            original_ra_model = model
                            model = 'RAH' + model[2:]  # Replace RA with RAH
                            # st.write(f"DEBUG - Converting RA to RAH: {original_ra_model} -> {model}")
                        
                        # Handle VOID models - convert "RAH0.5 VOID (+10%)" to "RAH0.5V"
                        if "VOID" in model.upper():
                            model = re.sub(r'(RAH\d+\.?\d*)\s+VOID.*', r'\1V', model, flags=re.IGNORECASE)
                        
                        # Handle STANDARD models - remove "STANDARD" from "RAH0.5 STANDARD"
                        if "STANDARD" in model.upper():
                            # Debug output for STANDARD models
                            # st.write(f"DEBUG - Found STANDARD in model: {model}")
                            # More flexible pattern - capture the base model and remove STANDARD
                            model = re.sub(r'(RAH\d+\.?\d*)\s+STANDARD.*', r'\1', model, flags=re.IGNORECASE)
                            # If that didn't work, try a simpler approach
                            if "STANDARD" in model.upper():
                                model = model.upper().replace("STANDARD", "").strip()
                            # st.write(f"DEBUG - After STANDARD removal: {model}")
                        
                        # Handle (PREM CONTROLS) suffix - remove from models like "RA4.0 (PREM CONTROLS)"
                        if "(PREM CONTROLS)" in model:
                            # st.write(f"DEBUG - Found PREM CONTROLS in model: {model}")
                            model = model.replace("(PREM CONTROLS)", "").strip()
                            # st.write(f"DEBUG - After PREM CONTROLS removal: {model}")
                        
                        # Debug model name conversion
                      
                    # Extract the numeric value from the orientation_volume string
                    ext_vol = None
                    if orientation_volume:
                        # Try to extract the numeric part
                        match = re.search(r'(\d+\.\d+|\d+)', str(orientation_volume))
                        if match:
                            ext_vol = match.group(1)
                    
                    model_data = {
                        'area_name': area_name,
                        'item_number': item_number,  # Include the item number from C column
                        'model': model,
                        'orientation_volume': orientation_volume,  # Store the full text for reference
                        'ext_vol': ext_vol,  # Store just the numeric part
                        'quantity': qty,
                        'width': width,
                        'length': length,
                        'height': height,
                        'location': location,  # Add location to the model data
                        'price': price_formatted  # Use the formatted price
                    }
                    
                    # Lookup additional specifications based on model
                    if model in RECOAIR_SPECS:
                        # Add p_drop, motor, and weight from lookup table
                        model_data['p_drop'] = RECOAIR_SPECS[model]['p_drop']
                        model_data['motor'] = RECOAIR_SPECS[model]['motor']
                        model_data['weight'] = RECOAIR_SPECS[model]['weight']
                        # st.write(f"DEBUG - Found specs for model {model}: P.DROP={model_data['p_drop']}, MOTOR={model_data['motor']}, WEIGHT={model_data['weight']}")
                    else:
                        # Use default values if model not found in specs
                        model_data['p_drop'] = 1050  # Default pressure drop
                        model_data['motor'] = ''     # Empty motor spec
                        model_data['weight'] = ''    # Empty weight
                        # st.write(f"DEBUG - No specs found for model {model}, using defaults")
                    
                    recoair_models.append(model_data)
                    project_data['recoair_models'].append(model_data)
                    
                    # Debug output
                    # st.write(f"DEBUG - Found RECOAIR model: {model}")
                    # st.write(f"DEBUG - Orientation/Volume: {orientation_volume}, Extracted volume: {ext_vol}")
                    # st.write(f"DEBUG - Details: W:{width} L:{length} H:{height} Price:{price}")
            
            # st.write(f"DEBUG - Total RECOAIR models found for {area_name}: {len(recoair_models)}")
            
           # st.write(k9_price)
           # st.write(n9_price)
            if n9_price > 0:
                project_data['recoair_control_data'][area_name] = {
                    'k9_price': k9_price,
                    'n9_price': n9_price,
                    'total_price': total_price  # Add total price to RECOAIR control data
                }
                project_data['has_recoair'] = True
               # st.write(project_data['global_recoair_n9_total'])
    
    # Calculate the total price from individual models
    recoair_subtotal = 0
    for model in project_data['recoair_models']:
        price = safe_convert_to_float(model.get('price', '0'))
        quantity = safe_convert_to_float(model.get('quantity', '1'))
        model_total = price * quantity
        recoair_subtotal += model_total
    
    # Add commissioning and delivery totals to the subtotal
    recoair_subtotal += project_data['global_recoair_delivery_total'] + project_data['global_recoair_commissioning_total']
    
    # Add 89.29 per RECOAIR sheet
    num_recoair_sheets = len([s for s in project_data.get('recoair_control_data', {}).keys()])
    if num_recoair_sheets > 0:
        # Add 89.29 for every RECOAIR sheet
        recoair_subtotal += (89.29 * num_recoair_sheets)
    
    # Calculate flat pack subtotal separately (don't add to recoair_subtotal)
    flat_pack_total = sum(item['flat_pack_price'] for item in project_data['recoair_flat_pack_details'])
    
    # Store the flat pack total for reference
    project_data['recoair_flat_pack_total'] = flat_pack_total
    project_data['recoair_flat_pack_formatted'] = f"{flat_pack_total:,.2f}"  # Format with commas and 2 decimal places
    
    # Also format each individual flat pack price in the details list
    for item in project_data['recoair_flat_pack_details']:
        item['flat_pack_price_formatted'] = f"{item['flat_pack_price']:,.2f}"  # Add formatted version
    
    # Set the subtotal values (excluding flat pack)
    project_data["recoair_subtotal"] = recoair_subtotal
    project_data["recoair_subtotal_formatted"] = f"{math.ceil(recoair_subtotal):,.2f}"
    
    # Calculate the total price (including flat pack)
    project_data['recoair_total_price'] = recoair_subtotal + flat_pack_total
    project_data['recoair_total_formatted'] = f"{math.ceil(recoair_subtotal + flat_pack_total):,.2f}"
    
    # Format the delivery and commissioning totals
    project_data['recoair_delivery_formatted'] = f"{math.ceil(project_data['global_recoair_delivery_total']):,.2f}"
    project_data['recoair_commissioning_formatted'] = f"{math.ceil(project_data['global_recoair_commissioning_total']):,.2f}"
    
    # st.write(f"DEBUG - RECOAIR subtotal: {project_data['recoair_subtotal_formatted']}")
    # st.write(f"DEBUG - RECOAIR total: {project_data['recoair_total_formatted']}")
    st.write(f"DEBUG - RECOAIR delivery total: {project_data['recoair_delivery_formatted']}")
    st.write(f"DEBUG - RECOAIR commissioning total: {project_data['recoair_commissioning_formatted']}")
    
    return project_data

def create_upload_section(col2, key_suffix=""):
    """Handle file upload and data extraction"""
    st.markdown("### Upload Existing Project")
    uploaded_file = st.file_uploader(
        "Upload Excel file to extract data", 
        type=['xlsx'],
        key=f"file_uploader_{key_suffix}"
    )
    
    if uploaded_file is not None:
        try:
            project_data = read_excel_file(uploaded_file)
            
            # Generate Word document from uploaded data
            if st.button("Generate Documents", key=f"generate_docs_{key_suffix}"):
                try:
                    # Get data from first sheet for Word doc
                    first_sheet = project_data['sheets'][0]
                    word_data = {
                        'Date': first_sheet['project_info']['date'],
                        'Project Number': first_sheet['project_info']['project_number'],
                        'Sales Contact': first_sheet['project_info']['sales_estimator'].split('/')[0],
                        'Estimator': first_sheet['project_info']['sales_estimator'],
                        'Estimator_Name': first_sheet['project_info']['estimator_name'],
                        'Estimator_Role': first_sheet['project_info']['estimator_role'],
                        'Customer': first_sheet['project_info']['customer'],
                        'Company': first_sheet['project_info']['company'],
                        'Address': first_sheet['project_info'].get('address', '')  # Add address from project_info
                    }
                    
                    # Generate Word document
                    word_doc_path = write_to_word_doc(word_data, project_data)
                    
                    # Save uploaded Excel temporarily to modify it
                    temp_excel_path = "temp_" + uploaded_file.name
                    with open(temp_excel_path, "wb") as f:
                        f.write(uploaded_file.getvalue())
                    
                    # Open the Excel file and check if it contains any SDU sheets
                    wb = openpyxl.load_workbook(temp_excel_path)
                    sdu_sheets = [sheet for sheet in wb.sheetnames if 'SDU' in sheet]
                    if sdu_sheets:
                        # Update project_data with the has_sdu flag
                        project_data['has_sdu'] = True
                        st.write("Found SDU sheets in the uploaded Excel file.")
                    
                    # Write totals and save
                    write_job_total(wb, project_data)
                    wb.save(temp_excel_path)
                    
                    # Create zip with Word doc and modified Excel
                    zip_path = create_download_zip(temp_excel_path, word_doc_path, project_data)
                    
                    # Show success message with balloons
                    st.balloons()
                    st.success("🎉 Documents generated successfully! 🎉")
                    
                    # Create download button for zip
                    with open(zip_path, "rb") as fp:
                        zip_contents = fp.read()
                        st.download_button(
                            label="⬇️ Download Documents",
                            data=zip_contents,
                            file_name="project_documents.zip",
                            mime="application/zip",
                            key=f"download_docs_{key_suffix}"
                        )
                    
                    # Clean up temp Excel file
                    os.remove(temp_excel_path)
                    
                except Exception as e:
                    st.error(f"Error generating documents: {str(e)}")
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")

def create_download_zip(excel_file, word_file, project_data=None):
    """Create a zip file containing both Excel and Word documents, and RECOAIR quotation if needed"""
    # Get project number and date from the Excel file
    wb = openpyxl.load_workbook(excel_file)
    project_number = None
    project_date = None
    
    # Look for project number and date in the first sheet that has 'CANOPY' in its title
    for sheet_name in wb.sheetnames:
        if 'CANOPY' in sheet_name:
            sheet = wb[sheet_name]
            # Get project number from cell C3
            project_number = sheet['C3'].value
            # Get date from cell G7
            project_date = sheet['G7'].value
            break
    
    if not project_number or not project_date:
        raise ValueError("Could not find project number or date in Excel file")
    
    # Format the date from DD/MM/YYYY to DD.MM.YYYY
    try:
        date_obj = datetime.strptime(project_date, "%d/%m/%Y")
        formatted_date = date_obj.strftime("%d.%m.%Y")
    except:
        formatted_date = project_date.replace('/', '.')
    
    # Create base filenames: one for Excel, one for Word documents
    base_filename_excel = f"{project_number} Cost Sheet {formatted_date}"
    base_filename_word = f"{project_number} Quotation {formatted_date}"
    
    # Create zip file
    zip_filename = f"{base_filename_excel}.zip" # Zip file can be named based on the Excel content
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        # Add Excel file with consistent naming
        zipf.write(excel_file, f"{base_filename_excel}.xlsx")
        
        # Add Word file with "Quotation" in the name
        if word_file: # word_file will be None if it wasn't generated (though current user code generates it always)
            zipf.write(word_file, f"{base_filename_word}.docx")
        
        # Check if RECOAIR data exists and add RECOAIR quotation if needed
        has_recoair = False
        has_sdu = False  # Add SDU flag
        
        # First check if RECOAIR sheets exist in the Excel file
        for sheet_name in wb.sheetnames:
            if 'RECOAIR' in sheet_name:
                has_recoair = True
            if 'SDU' in sheet_name:
                has_sdu = True
        
        # Also check if project_data contains RECOAIR data
        if not has_recoair and project_data and 'has_recoair' in project_data and project_data['has_recoair']:
            has_recoair = True
            
        # Check if project_data contains SDU flag
        if not has_sdu and project_data and 'has_sdu' in project_data and project_data['has_sdu']:
            has_sdu = True
        
        # If RECOAIR is used, include the RECOAIR quotation
        if has_recoair:
            recoair_doc_path = "resources/Halton RECO Quotation Jan 2025 (2).docx"
            
            # Check if the RECOAIR document exists
            if os.path.exists(recoair_doc_path):
                # Create a copy of the RECOAIR document to customize
                custom_recoair_doc = f"temp_recoair_{project_number}.docx"
                shutil.copy(recoair_doc_path, custom_recoair_doc)
                
                # If project_data is available, we can customize the RECOAIR document
                if project_data:
                    try:
                        # Load the template
                        doc = DocxTemplate(custom_recoair_doc)
                        
                        # Get first sheet data for project info
                        first_sheet = project_data['sheets'][0]
                        project_info = first_sheet['project_info']
                        
                        # Prepare context data for the RECOAIR document (same as for main document)
                        # Get address safely from project_data
                        address = ''
                        if 'sheets' in project_data and project_data['sheets']:
                            first_sheet = project_data['sheets'][0]
                            if 'project_info' in first_sheet:
                                address = first_sheet['project_info'].get('address', '')
                        
                        # Get sales contact name and number
                        sales_initials = project_info.get('sales_estimator', '').split('/')[0]
                        
                        # Look up full sales contact name from initials
                        sales_contact_name = next((name for name in contacts.keys() if name.startswith(sales_initials)), None)
                        
                        # If not found by prefix, try to match by initials
                        if not sales_contact_name:
                            for full_name in contacts.keys():
                                if get_initials(full_name) == sales_initials:
                                    sales_contact_name = full_name
                                    break
                                    
                        # If still not found, use the initials
                        if not sales_contact_name:
                            st.warning(f"Could not find contact information for {sales_initials}")
                            sales_contact_name = sales_initials
                            contact_number = ""
                        else:
                            contact_number = contacts.get(sales_contact_name, "")
                        
                        # Extract customer's first name
                        customer_full = project_info.get('customer', '')
                        customer_first_name = customer_full.split()[0] if customer_full else ""
                        
                        # Get estimator information
                        estimator_initials = project_info.get('sales_estimator', '').split('/')[-1] if '/' in project_info.get('sales_estimator', '') else ''
                        
                        # Find estimator name and role from initials
                        estimator_name = ''
                        estimator_role = ''
                        for full_name, role in estimators.items():
                            if get_initials(full_name) == estimator_initials:
                                estimator_name = full_name
                                estimator_role = role
                                break
                                
                        # If not found, use values from project_info if available
                        if not estimator_name and 'estimator_name' in project_info:
                            estimator_name = project_info.get('estimator_name', '')
                        if not estimator_role and 'estimator_role' in project_info:
                            estimator_role = project_info.get('estimator_role', '')
                        
                        # Get current revision
                        current_revision = first_sheet.get('revision', 'A')
                        
                        # Create Halton reference with full sales contact initials
                        halton_ref = f"{project_info.get('project_number', '')}/{get_initials(sales_contact_name)}/{estimator_initials}/{current_revision}" if current_revision != 'A' else f"{project_info.get('project_number', '')}/{get_initials(sales_contact_name)}/{estimator_initials}"
                        
                        # Create context data
                        context = {
                            # Customer and company information
                            'customer': project_info.get('customer', ''),
                            'customer_first_name': customer_first_name,
                            'company': project_info.get('company', ''),
                            'address': address,  # Lowercase for backward compatibility 
                            'Address': address,  # Uppercase A for template compatibility
                            
                            # Project information
                            'date': project_info.get('date', ''),
                            'project_number': halton_ref,
                            'project_name': project_info.get('project_name', ''),
                            'location': project_info.get('location', ''),
                            'quote_title': "QUOTATION" if current_revision == 'A' else f"QUOTATION - Revision {current_revision}",
                            
                            # Contact information
                            'sales_contact_name': sales_contact_name,
                            'contact_number': contact_number,
                            'estimator_name': estimator_name,
                            'estimator_role': estimator_role,
                            
                            # Additional data for consistency with main document
                            'revision': current_revision,
                            'estimator': estimator_initials,  # Use initials for estimator field
                            'sales_contact': sales_contact_name,  # Use full name instead of initials
                            'Project Number': project_info.get('project_number', ''),
                            'Project Name': project_info.get('project_name', ''),
                            'Location': project_info.get('location', ''),
                            'Date': project_info.get('date', ''),
                            'Sales Contact': sales_contact_name,  # Use full name instead of initials
                            'Customer': project_info.get('customer', ''),
                            'Company': project_info.get('company', ''),
                            
                            # Additional data
                            'areas': [],
                            'has_recoair': True,
                            'has_sdu': has_sdu,  # Add SDU flag to RECOAIR context
                            
                            # Add delivery and commissioning totals
                            'recoair_delivery_total': project_data.get('recoair_delivery_formatted', '0.00'),
                            'recoair_commissioning_total': project_data.get('recoair_commissioning_formatted', '0.00'),
                            
                            # Add RECOAIR subtotal (equipment only, excluding flat pack)
                            'recoair_subtotal': project_data.get('recoair_subtotal_formatted', '0.00'),
                            
                            # Add flat pack details
                            'recoair_flat_pack_details': project_data.get('recoair_flat_pack_details', []),
                            'recoair_flat_pack_total': project_data.get('recoair_flat_pack_formatted', '0.00'),
                            
                            # Add final total (equipment + flat pack)
                            'recoair_grand_total': project_data.get('recoair_total_formatted', '0.00')
                        }
                        # st.write(context)
                        # Add areas with RECOAIR data
                        if 'recoair_control_data' in project_data:
                            for area_name, recoair_data in project_data['recoair_control_data'].items():
                                context['areas'].append({
                                    'name': area_name,
                                    'has_recoair': True,
                                    'recoair_price': format_price(recoair_data.get('n9_price', 0)),
                                    'recoair_total': format_price(recoair_data.get('total_price', 0))
                                })
                        
                        # Add RECOAIR models data for the table
                        if 'recoair_models' in project_data and project_data['recoair_models']:
                            # Ensure each model has an item number
                            for model in project_data['recoair_models']:
                                if 'item_number' not in model:
                                    model['item_number'] = "1.01"  # Default if missing
                            
                            context['recoair_models'] = project_data['recoair_models']
                            # Debug output
                            st.write(f"DEBUG - Adding {len(project_data['recoair_models'])} RECOAIR models to the document template")
                        else:
                            context['recoair_models'] = []
                        
                        # Add RECOAIR subtotal and total prices
                        context['recoair_subtotal'] = project_data.get('recoair_subtotal_formatted', '0.00')
                        context['recoair_total'] = project_data.get('recoair_total_formatted', '0.00')
                        st.write(f"DEBUG - RECOAIR subtotal in context: {context['recoair_subtotal']}")
                        st.write(f"DEBUG - RECOAIR total in context: {context['recoair_total']}")
                        
                        # Render the document
                        doc.render(context)
                        doc.save(custom_recoair_doc)
                        
                        # Add the customized RECOAIR document to the zip with "Quotation" in the name
                        zipf.write(custom_recoair_doc, f"{base_filename_word} - RECOAIR.docx")
                        
                        # Clean up temporary file
                        os.remove(custom_recoair_doc)
                    except Exception as e:
                        # If customization fails, use the original template
                        st.warning(f"Could not customize RECOAIR document: {str(e)}")
                        zipf.write(recoair_doc_path, f"{base_filename_word} - RECOAIR.docx")
                else:
                    # If no project_data, just include the original template
                    zipf.write(recoair_doc_path, f"{base_filename_word} - RECOAIR.docx")
            else:
                st.warning(f"RECOAIR document template not found at: {recoair_doc_path}")
    
    return zip_filename

def write_job_total(workbook, project_data):
    """Write total prices and project info to JOB TOTAL sheet"""
    if 'JOB TOTAL' in workbook.sheetnames:
        total_sheet = workbook['JOB TOTAL']
        
        # Get first sheet data for project info
        first_sheet = project_data['sheets'][0]
        project_info = first_sheet['project_info']
        
        # Write project information
        total_sheet['C3'] = project_info['project_number']
        customer_company = f"{project_info['customer']} ({project_info['company']})" if project_info['company'] else project_info['customer']
        total_sheet['C5'] = customer_company
        total_sheet['C7'] = project_info['sales_estimator']
        
        total_sheet['G3'] = project_info['project_name']
        total_sheet['G5'] = project_info['location']
        total_sheet['G7'] = project_info['date']
        
        # Get revision from first used sheet
        revision = next(sheet['revision'] for sheet in project_data['sheets']
                       if any(canopy['reference_number'] != 'ITEM' 
                             for canopy in sheet['canopies']))
        total_sheet['O7'] = revision
        
        # Initialize totals
        job_total = 0.0
        k9_total = 0.0
        
       # st.write("\n🔄 Processing sheets for totals:")
        # Process each sheet
        for sheet_data in project_data['sheets']:
           # st.write(f"\n📄 Processing sheet: {sheet_data.get('sheet_name', 'Unknown')}")
            # Only add to total if sheet has canopies (is used)
            if any(canopy['reference_number'] != 'ITEM' and 
                  canopy['model'] != 'CANOPY TYPE' and
                  canopy['reference_number'] != 'DELIVERY & INSTALLATION'
                  for canopy in sheet_data['canopies']):
                
                # Convert string prices to numbers
                try:
                    sheet_total = float(str(sheet_data['total_price']).replace(',', ''))
                    sheet_k9 = float(str(sheet_data['k9_total']).replace(',', ''))
                except (ValueError, TypeError):
                    sheet_total = 0.0
                    sheet_k9 = 0.0
                
                # st.write(f"💰 Total price from sheet: {sheet_total}")
                # st.write(f"💰 K9 total from sheet: {sheet_k9}")
                
                job_total += sheet_total
                k9_total += sheet_k9
        
        # Convert fire suppression totals to numbers
        try:
            fs_k9_total = float(str(project_data['global_fs_k9_total']).replace(',', ''))
            fs_n9_total = float(str(project_data['global_fs_n9_total']).replace(',', ''))
            #st.write(fs_k9_total)
            uv_k9_total = float(str(project_data['global_uv_k9_total']).replace(',', ''))
            uv_n9_total = float(str(project_data['global_uv_n9_total']).replace(',', ''))
            recoair_k9_total = float(str(project_data.get('global_recoair_k9_total', 0)).replace(',', ''))
            recoair_n9_total = float(str(project_data.get('global_recoair_n9_total', 0)).replace(',', ''))
            # st.write(project_data)
        except (ValueError, TypeError):
            fs_k9_total = 0.0
            fs_n9_total = 0.0
            uv_k9_total = 0.0
            uv_n9_total = 0.0
            recoair_k9_total = 0.0
            recoair_n9_total = 0.0
        
        # Debug totals for JOB TOTAL sheet
        # st.write("### Values Being Written to JOB TOTAL Sheet")
        # st.write(f"Job Total (T16): £{job_total}")
        # st.write(f"K9 Total (S16): £{k9_total}")
        # st.write(f"Fire Suppression - Cost (S17): £{fs_k9_total}")
        # st.write(f"Fire Suppression - Selling Price (T17): £{fs_n9_total}")
        # st.write(f"UV-C - Cost (S21): £{uv_k9_total}")
        # st.write(f"UV-C - Selling Price (T21): £{uv_n9_total}")
        # st.write(f"RECOAIR - Cost (S24): £{recoair_k9_total}")
        # st.write(f"RECOAIR - Selling Price (T24): £{recoair_n9_total}")
        
        # Write the totals as numbers
        total_sheet['T16'] = job_total  # Main job total
        total_sheet['S16'] = k9_total   # Main K9 total
        total_sheet['S17'] = fs_k9_total  # Fire suppression K9 total
        total_sheet['T17'] = fs_n9_total  # Fire suppression N9 total
        
        # Only write UV totals if there is UV data
        has_uv_data = bool(project_data.get('uv_control_data', {}))
        if has_uv_data:
            total_sheet['S21'] = uv_k9_total  # UV-C K9 total
            total_sheet['T21'] = uv_n9_total  # UV-C N9 total
        else:
            # Zero out the cells if no UV data
            total_sheet['S21'] = 0  # UV-C K9 total
            total_sheet['T21'] = 0  # UV-C N9 total
            
        total_sheet['S24'] = recoair_k9_total  # RECOAIR K9 total (cost)
        total_sheet['T24'] = recoair_n9_total  # RECOAIR N9 total (selling price)
        
        # Check if RECOAIR sheets are included in the job_total calculation
        recoair_already_included = False
        for sheet_data in project_data['sheets']:
            if 'sheet_name' in sheet_data and 'RECOAIR' in str(sheet_data['sheet_name']):
                recoair_already_included = True
                break
        
        # Calculate the correct grand total, avoiding double-counting
        # Only add UV totals if there is actual UV data
        if recoair_already_included:
            # If RECOAIR sheets are already processed in job_total, don't add recoair_n9_total again
            grand_total = job_total + fs_n9_total + (uv_n9_total if has_uv_data else 0)
        else:
            # If RECOAIR is not part of job_total, include it
            grand_total = job_total + fs_n9_total + (uv_n9_total if has_uv_data else 0) + recoair_n9_total
        
        # Override the formula cells for subtotal and grand total in the template
        # Based on your image, row 27 contains subtotal and row 28 contains the grand total
        total_sheet['T27'] = grand_total  # Override subtotal formula
        total_sheet['T28'] = grand_total  # Override grand total formula
        
        # Ensure Lists sheet is hidden here as well
        if 'Lists' in workbook.sheetnames:
            workbook['Lists'].sheet_state = 'hidden'
    
    # No return value needed

def write_to_recoair_doc(data, project_data, output_path="recoair_output.docx"):
    """Generate a RECOAIR quotation document using the provided data"""
    try:
        # Check if RECOAIR data exists
        has_recoair = False
        
        # Check if project_data has RECOAIR flag
        if project_data and 'has_recoair' in project_data and project_data['has_recoair']:
            has_recoair = True
        
        # Or if there's recoair_control_data with entries
        if not has_recoair and project_data and 'recoair_control_data' in project_data and project_data['recoair_control_data']:
            has_recoair = True
            
        # If no RECOAIR data, don't generate document
        if not has_recoair:
            return None
            
        # Load the RECOAIR template
        recoair_doc_path = "resources/Halton RECO Quotation Jan 2025 (2).docx"
        if not os.path.exists(recoair_doc_path):
            st.warning(f"RECOAIR document template not found at: {recoair_doc_path}")
            return None
            
        # Load the template
        doc = DocxTemplate(recoair_doc_path)
        
        # Get first sheet data for project info
        first_sheet = project_data['sheets'][0]
        project_info = first_sheet['project_info']
        
        # Get address from project_data
        address = ''
        if 'sheets' in project_data and project_data['sheets']:
            first_sheet = project_data['sheets'][0]
            if 'project_info' in first_sheet:
                address = first_sheet['project_info'].get('address', '')
        
        # Get sales contact name and number
        sales_initials = data.get('Sales Contact', '')
        sales_contact_name = next((name for name in contacts.keys() if name.startswith(sales_initials)), None)
        
        # If no match found, try to find by initials
        if not sales_contact_name:
            for full_name in contacts.keys():
                if get_initials(full_name) == sales_initials:
                    sales_contact_name = full_name
                    break
        
        # If still not found, use the initials
        if not sales_contact_name:
            st.error(f"Could not find contact information for {sales_initials}")
            sales_contact_name = sales_initials
            contact_number = ""
        else:
            contact_number = contacts[sales_contact_name]
        
        # Extract customer's first name
        customer_full = data['Customer']
        customer_first_name = customer_full.split()[0] if customer_full else ""
        
        # Get estimator initials from the sales_estimator field
        estimator_initials = data.get('Estimator', '').split('/')[-1] if '/' in data.get('Estimator', '') else ''
        
        # Find estimator name and role from initials
        estimator_name = ''
        estimator_role = ''
        for full_name, role in estimators.items():
            if get_initials(full_name) == estimator_initials:
                estimator_name = full_name
                estimator_role = role
                break
        
        # Get current revision
        current_revision = project_data['sheets'][0].get('revision', 'A')
        
        # Create Halton reference
        halton_ref = f"{data['Project Number']}/{get_initials(sales_contact_name)}/{estimator_initials}/{current_revision}" if current_revision != 'A' else f"{data['Project Number']}/{get_initials(sales_contact_name)}/{estimator_initials}"
        
        # Create quote title based on revision
        quote_title = "QUOTATION" if current_revision == 'A' else f"QUOTATION - Revision {current_revision}"
        
        # Create context data
        context = {
            # Customer and company information
            'customer': data['Customer'],
            'customer_first_name': customer_first_name,
            'company': data.get('Company', ''),
            'address': address,  # Lowercase for backward compatibility 
            'Address': address,  # Uppercase A for template compatibility
            
            # Project information
            'date': data.get('Date', ''),
            'project_number': halton_ref,
            'project_name': project_info.get('project_name', ''),
            'location': project_info.get('location', ''),
            'quote_title': quote_title,
            
            # Contact information
            'sales_contact_name': sales_contact_name,
            'contact_number': contact_number,
            'estimator_name': estimator_name,
            'estimator_role': estimator_role,
            
            # Additional data for consistency with main document
            'revision': current_revision,
            'estimator': data.get('Estimator', ''),
            'sales_contact': data.get('Sales Contact', ''),
            'Project Number': data.get('Project Number', ''),
            'Project Name': project_info.get('project_name', ''),
            'Location': project_info.get('location', ''),
            'Date': data.get('Date', ''),
            'Sales Contact': data.get('Sales Contact', ''),
            'Cu∫stomer': data.get('Customer', ''),
            'Company': data.get('Company', ''),
            
            # Additional data
            'areas': [],
            'has_recoair': True,
            
            # Add delivery and commissioning totals
            'recoair_delivery_total': project_data.get('recoair_delivery_formatted', '0.00'),
            'recoair_commissioning_total': project_data.get('recoair_commissioning_formatted', '0.00'),
            
            # Add RECOAIR subtotal (equipment only, excluding flat pack)
            'recoair_subtotal': project_data.get('recoair_subtotal_formatted', '0.00'),
            
            # Add flat pack details
            'recoair_flat_pack_details': project_data.get('recoair_flat_pack_details', []),
            'recoair_flat_pack_total': project_data.get('recoair_flat_pack_formatted', '0.00'),
            
            # Add final total (equipment + flat pack)
            'recoair_grand_total': project_data.get('recoair_total_formatted', '0.00')
        }
        st.write(context)
        # Add areas with RECOAIR data
        if 'recoair_control_data' in project_data:
            for area_name, recoair_data in project_data['recoair_control_data'].items():
                context['areas'].append({
                    'name': area_name,
                    'has_recoair': True,
                    'recoair_price': format_price(recoair_data.get('n9_price', 0)),
                    'recoair_total': format_price(recoair_data.get('total_price', 0))
                })
        
        # Add RECOAIR models data for the table
        if 'recoair_models' in project_data and project_data['recoair_models']:
            context['recoair_models'] = project_data['recoair_models']
            # Debug output
            st.write(f"DEBUG - Adding {len(project_data['recoair_models'])} RECOAIR models to the document template")
        else:
            context['recoair_models'] = []
        
        # Render the document
        doc.render(context)
        
        # Save the document
        doc.save(output_path)
        return output_path
    except Exception as e:
        st.error(f"Error generating RECOAIR document: {str(e)}")
        return None

def create_revision_tab():
    st.title("📝 Revise Cost Sheet")
    
    # Only show the Upload Original File section - the upload project section is now in its own tab
    st.subheader("Upload Original File")
    uploaded_file = st.file_uploader(
        "Upload Excel file to revise", 
        type=['xlsx'],
        key="revision_tab_uploader"
    )
    
    if uploaded_file is not None:
        try:
            # Read the Excel file
            project_data = read_excel_file(uploaded_file)
            
            # Get project name and number from first sheet
            first_sheet = project_data['sheets'][0]
            project_name = first_sheet['project_info']['project_name']
            project_number = first_sheet['project_info']['project_number']
            current_revision = first_sheet['revision']
            
            # Add new floor/area section
            st.markdown("---")
            st.subheader("Add New Floor/Area")
            new_floor = st.text_input("New Floor Name (e.g., Ground Floor)")
            new_area = st.text_input("New Area Name (e.g., Kitchen)")
            
            if st.button("Add Floor/Area", key="add_floor_area_button") and new_floor and new_area:
                try:
                    new_filename = add_new_floor_area(uploaded_file, new_floor, new_area, current_revision)
                    st.success(f"Added new floor/area: {new_floor} - {new_area}")
                    
                    # Add download button for the updated file
                    with open(new_filename, "rb") as file:
                        st.download_button(
                            label="Download Updated Excel",
                            data=file,
                            file_name=os.path.basename(new_filename),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_updated_excel_button"
                        )
                except Exception as e:
                    st.error(f"Error adding floor/area: {str(e)}")
            
            # Create new revision section
            st.markdown("---")
            st.subheader("Create New Revision")
            if st.button("Create New Revision", key="create_revision_button"):
                next_rev = chr(ord(current_revision) + 1)
                try:
                    new_filename = create_new_revision(uploaded_file, project_name, project_number, next_rev)
                    st.success(f"Created Revision {next_rev}!")
                    
                    # Add download button for the new revision
                    with open(new_filename, "rb") as file:
                        st.download_button(
                            label=f"Download Revision {next_rev}",
                            data=file,
                            file_name=os.path.basename(new_filename),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_revision_button"
                        )
                except Exception as e:
                    st.error(f"Error creating revision: {str(e)}")
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")

def create_new_revision(uploaded_file, project_name, project_number, next_rev):
    """Create a new revision of the Excel file"""
    # Create temporary file
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    # Open workbook
    wb = openpyxl.load_workbook(temp_excel_path)
    
    # Update revision for each sheet and add dropdowns
    for sheet_name in [s for s in wb.sheetnames if 'CANOPY' in s or s == 'JOB TOTAL']:
        sheet = wb[sheet_name]
        sheet['O7'] = next_rev
        
        # Add dropdowns to CANOPY sheets
        if 'CANOPY' in sheet_name:
            add_dropdowns_to_sheet(wb, sheet, 12)
    
    # Create folder name with project details
    folder_name = f"{project_name} - {project_number} (Revision {next_rev})"
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    
    # Save as new file in revision folder
    new_filename = os.path.join(folder_name, uploaded_file.name.replace('.xlsx', f'_Rev{next_rev}.xlsx'))
    wb.save(new_filename)
    
    # Clean up temp file
    os.remove(temp_excel_path)
    
    return new_filename

def add_new_floor_area(uploaded_file, new_floor, new_area, current_revision):
    """Add a new floor/area sheet to the Excel file"""
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    wb = openpyxl.load_workbook(temp_excel_path)
    
    # Find first empty CANOPY sheet
    canopy_sheets = [s for s in wb.sheetnames if 'CANOPY' in s]
    new_sheet = None
    for sheet_name in canopy_sheets:
        sheet = wb[sheet_name]
        if not sheet['B12'].value:  # Check if sheet is empty (no canopies)
            new_sheet = sheet
            break
    
    if new_sheet:
        # Update sheet name and title
        new_sheet_name = f"CANOPY - {new_floor} ({len(canopy_sheets) + 1})"
        new_sheet.title = new_sheet_name
        new_sheet['B1'] = f"{new_floor} - {new_area}"
        new_sheet['O7'] = current_revision
        
        # Add dropdowns to the new sheet
        add_dropdowns_to_sheet(wb, new_sheet, 12)
        
        # Save workbook
        new_filename = uploaded_file.name.replace('.xlsx', '_updated.xlsx')
        wb.save(new_filename)
        os.remove(temp_excel_path)
        return new_filename
    else:
        os.remove(temp_excel_path)
        raise Exception("No empty CANOPY sheets available")

def edit_floor_area_name(uploaded_file, selected_sheet, new_name, current_revision):
    """Edit the name of an existing floor/area"""
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    wb = openpyxl.load_workbook(temp_excel_path)
    sheet = wb[selected_sheet]
    
    # Update sheet title and name
    sheet['B1'] = new_name
    sheet.title = f"CANOPY - {new_name}"
    
    # Save workbook
    new_filename = uploaded_file.name.replace('.xlsx', '_updated.xlsx')
    wb.save(new_filename)
    os.remove(temp_excel_path)
    return new_filename

def add_new_canopy(uploaded_file, selected_sheet, ref_number, model, config, current_revision):
    """Add a new canopy to an existing area"""
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    wb = openpyxl.load_workbook(temp_excel_path)
    sheet = wb[selected_sheet]
    
    # Find first empty canopy slot
    row = 12
    while row <= 181:
        if not sheet[f'B{row}'].value:
            # Write canopy data
            sheet[f'B{row}'] = ref_number
            sheet[f'C{row + 2}'] = config
            sheet[f'D{row + 2}'] = model
            
            # Add standard entries
            sheet[f'C{row + 3}'] = "LIGHT SELECTION"
            sheet[f'C{row + 4}'] = "SELECT WORKS"
            sheet[f'C{row + 5}'] = "SELECT WORKS"
            sheet[f'C{row + 6}'] = "BIM/ REVIT per CANOPY"
            sheet[f'D{row + 6}'] = "1"
            break
        row += 17
    
    # Save workbook
    new_filename = uploaded_file.name.replace('.xlsx', '_updated.xlsx')
    wb.save(new_filename)
    os.remove(temp_excel_path)
    return new_filename

def edit_canopy(uploaded_file, selected_sheet, selected_canopy, new_model, new_config, current_revision):
    """Edit an existing canopy's properties"""
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    wb = openpyxl.load_workbook(temp_excel_path)
    sheet = wb[selected_sheet]
    
    # Find the canopy
    row = 12
    while row <= 181:
        if sheet[f'B{row}'].value == selected_canopy:
            # Update canopy data
            sheet[f'C{row + 2}'] = new_config
            sheet[f'D{row + 2}'] = new_model
            break
        row += 17
    
    # Save workbook
    new_filename = uploaded_file.name.replace('.xlsx', '_updated.xlsx')
    wb.save(new_filename)
    os.remove(temp_excel_path)
    return new_filename

def update_cladding(uploaded_file, selected_sheet, selected_canopy, width, height, positions, current_revision):
    """Update wall cladding for a canopy"""
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    wb = openpyxl.load_workbook(temp_excel_path)
    sheet = wb[selected_sheet]
    
    # Find the canopy
    row = 12
    while row <= 181:
        if sheet[f'B{row}'].value == selected_canopy:
            cladding_row = row + 7
            # Update cladding data
            sheet[f'C{cladding_row}'] = "2M² (HFL)"
            sheet[f'Q{cladding_row}'] = width
            sheet[f'R{cladding_row}'] = height
            sheet[f'S{cladding_row}'] = ','.join(positions)
            
            # Create formatted display string
            positions_str = '/'.join(positions)
            cladding_info = f"2M² (HFL) - {width}x{height}mm ({positions_str})"
            sheet[f'T{cladding_row}'] = cladding_info
            break
        row += 17
    
    # Save workbook
    new_filename = uploaded_file.name.replace('.xlsx', '_updated.xlsx')
    wb.save(new_filename)
    os.remove(temp_excel_path)
    return new_filename

def delete_canopy(uploaded_file, selected_sheet, selected_canopy, current_revision):
    """Delete a canopy from an area"""
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    wb = openpyxl.load_workbook(temp_excel_path)
    sheet = wb[selected_sheet]
    
    # Find the canopy
    row = 12
    while row <= 181:
        if sheet[f'B{row}'].value == selected_canopy:
            # Clear canopy data (17 rows)
            for r in range(row, row + 17):
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
                    sheet[f'{col}{r}'].value = None
            break
        row += 17
    
    # Save workbook
    new_filename = uploaded_file.name.replace('.xlsx', '_updated.xlsx')
    wb.save(new_filename)
    os.remove(temp_excel_path)
    return new_filename

def reorder_canopies(uploaded_file, selected_sheet, new_order, current_revision):
    """Reorder canopies in an area"""
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    wb = openpyxl.load_workbook(temp_excel_path)
    sheet = wb[selected_sheet]
    
    # Store current canopy data
    canopy_data = {}
    row = 12
    while row <= 181:
        ref = sheet[f'B{row}'].value
        if ref and ref != 'ITEM':
            # Store all values for this canopy section (17 rows)
            canopy_data[ref] = {}
            for r in range(row, row + 17):
                canopy_data[ref][r-row] = {}
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
                    canopy_data[ref][r-row][col] = sheet[f'{col}{r}'].value
        row += 17
    
    # Clear all canopy data
    row = 12
    while row <= 181:
        for r in range(row, row + 17):
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
                sheet[f'{col}{r}'].value = None
        row += 17
    
    # Write canopies in new order
    row = 12
    for ref in new_order:
        if ref in canopy_data:
            for r_offset, col_data in canopy_data[ref].items():
                for col, value in col_data.items():
                    sheet[f'{col}{row + r_offset}'].value = value
        row += 17
    
    # Save workbook
    new_filename = uploaded_file.name.replace('.xlsx', '_updated.xlsx')
    wb.save(new_filename)
    os.remove(temp_excel_path)
    return new_filename

def copy_area_to_new_floor(uploaded_file, source_sheet, new_floor, new_area, current_revision):
    """Copy an area to a new floor"""
    temp_excel_path = "temp_" + uploaded_file.name
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    
    wb = openpyxl.load_workbook(temp_excel_path)
    source = wb[source_sheet]
    
    # Find first empty CANOPY sheet
    canopy_sheets = [s for s in wb.sheetnames if 'CANOPY' in s]
    target = None
    for sheet_name in canopy_sheets:
        sheet = wb[sheet_name]
        if not sheet['B12'].value:  # Check if sheet is empty
            target = sheet
            break
    
    if target:
        # Copy all cells from source to target
        for row in source.rows:
            for cell in row:
                target[cell.coordinate].value = cell.value
        
        # Update sheet name and title
        target.title = f"CANOPY - {new_floor} ({len(canopy_sheets) + 1})"
        target['B1'] = f"{new_floor} - {new_area}"
        target['O7'] = current_revision
        
        # Save workbook
        new_filename = uploaded_file.name.replace('.xlsx', '_updated.xlsx')
        wb.save(new_filename)
        os.remove(temp_excel_path)
        return new_filename
    else:
        os.remove(temp_excel_path)
        raise Exception("No empty CANOPY sheets available")

def add_fire_suppression_dropdown(sheet):
    """Add dropdown validation for fire suppression options"""
    if 'Lists' not in sheet.parent.sheetnames:
        list_sheet = sheet.parent.create_sheet('Lists')
    else:
        list_sheet = sheet.parent['Lists']
    
    # Add fire suppression options to Lists sheet (column H)
    fire_suppression_options = [
        "SELECT TANK SYSTEM",
        "1 TANK SYSTEM",
        "1 TANK TRAVEL HUB",
        "1 TANK DISTANCE",
        "NOBEL",
        "AMAREX",
        "OTHER",
        "2 TANK SYSTEM",
        "2 TANK TRAVEL HUB",
        "2 TANK DISTANCE",
        "3 TANK SYSTEM",
        "3 TANK TRAVEL HUB",
        "3 TANK DISTANCE",
        "4 TANK SYSTEM",
        "4 TANK TRAVEL HUB",
        "4 TANK DISTANCE",
        "5 TANK SYSTEM",
        "5 TANK TRAVEL HUB",
        "5 TANK DISTANCE",
        "6 TANK SYSTEM",
        "6 TANK TRAVEL HUB",
        "6 TANK DISTANCE"
    ]
    
    # Add tank installation options to Lists sheet (column I)
    tank_install_options = [
        "1 TANK",
        "1 TANK DISTANCE",
        "2 TANK",
        "2 TANK DISTANCE",
        "3 TANK",
        "3 TANK DISTANCE",
        "4 TANK",
        "4 TANK DISTANCE",
        "5 TANK",
        "5 TANK DISTANCE",
        "6 TANK",
        "6 TANK DISTANCE"
    ]
    
    # Write fire suppression options to Lists sheet
    for i, option in enumerate(fire_suppression_options, 1):
        list_sheet[f'H{i}'] = option
        
    # Write tank installation options to Lists sheet
    for i, option in enumerate(tank_install_options, 1):
        list_sheet[f'I{i}'] = option
    
    # Create validation for fire suppression
    dv_fs = DataValidation(
        type="list",
        formula1=f"Lists!$H$1:$H${len(fire_suppression_options)}",
        allow_blank=True
    )
    sheet.add_data_validation(dv_fs)
    
    # Create validation for tank installation
    dv_tank = DataValidation(
        type="list",
        formula1=f"Lists!$I$1:$I${len(tank_install_options)}",
        allow_blank=True
    )
    sheet.add_data_validation(dv_tank)
    
    # Add fire suppression validation to C16 and every 17 rows after
    row = 16
    while row <= sheet.max_row:
        dv_fs.add(f"C{row}")
        row += 17
    
    # Add tank installation validation to C17 and every 17 rows after
    row = 17
    while row <= sheet.max_row:
        dv_tank.add(f"C{row}")
        row += 17

def create_upload_project_tab():
    """Create the Upload Project tab with functionality to upload and process existing projects"""
    st.title("📤 Upload Existing Project")
    
    # Use the create_upload_section function directly in this tab
    # We don't need columns since this tab is dedicated to just the upload functionality
    create_upload_section(st, "upload_tab")

def add_recoair_internal_external_dropdowns(sheet):
    """Add dropdowns to RECOAIR sheets:
    1. Internal/External options to cells I14-I28 and T14-T28
    2. Plant hire options to cells E38 and E39
    """
    try:
        # Get or create the Lists sheet
        if 'Lists' not in sheet.parent.sheetnames:
            list_sheet = sheet.parent.create_sheet('Lists')
        else:
            list_sheet = sheet.parent['Lists']
        
        # Add column headers
        sheet['C13'] = "Model"        # Model column header
        sheet['D13'] = "Orientation"  # Orientation column header
        sheet['I13'] = "Location"     # Location column header
        sheet['T13'] = "Location"     # Second location column header
        
        # --- LOCATION DROPDOWN (I14-I28 and T14-T28) ---
        # Define options
        location_options = ["Internal", "External"]
        
        # Clear any existing entries
        for i in range(1, 10):  # Clear enough rows for our options
            if list_sheet.cell(row=i, column=10).value:  # Column J is 10
                list_sheet.cell(row=i, column=10).value = None
        
        # Write options to Lists sheet in column J
        for i, option in enumerate(location_options, 1):
            list_sheet.cell(row=i, column=10).value = option  # Column J is 10
        
        # Create data validation (directly using string formula for better compatibility)
        location_dv = DataValidation(
            type="list", 
            formula1="Lists!$J$1:$J$2",  # Explicitly reference the exact range
            allow_blank=True
        )
        
        # Add the data validation to the sheet
        sheet.add_data_validation(location_dv)
        
        # Add column headers
        sheet['I13'] = "Location"
        sheet['T13'] = "Location"
        
        # Apply to cells I14 through I28 AND T14 through T28
        for row in range(14, 29):
            cell_i = f'I{row}'
            cell_t = f'T{row}'
            
            location_dv.add(cell_i)
            location_dv.add(cell_t)
            
            # Don't set default values, just make cells empty to show dropdown arrows
            sheet[cell_i].value = None
            sheet[cell_t].value = None
            
        # --- PLANT HIRE DROPDOWN (E38 and E39) ---
        # Define plant hire options
        plant_hire_options = [
            "", "SL10 GENIE", "EXTENSION FORKS", "2.5M COMBI LADDER",
            "1.5M PODIUM", "3M TOWER", "COMBI LADDER", "PECO LIFT",
            "3M YOUNGMAN BOARD", "GS1930 SCISSOR LIFT",
            "4-6 SHERASCOPIC", "7-9 SHERASCOPIC"
        ]
        
        # Clear any existing entries in column K
        for i in range(1, 20):  # Clear enough rows for our options
            if list_sheet.cell(row=i, column=11).value:  # Column K is 11
                list_sheet.cell(row=i, column=11).value = None
        
        # Write options to Lists sheet in column K
        for i, option in enumerate(plant_hire_options, 1):
            list_sheet.cell(row=i, column=11).value = option  # Column K is 11
        
        # Create data validation for plant hire
        plant_hire_dv = DataValidation(
            type="list", 
            formula1=f"Lists!$K$1:$K${len(plant_hire_options)}",
            allow_blank=True
        )
        
        # Add the data validation to the sheet
        sheet.add_data_validation(plant_hire_dv)
        
        # Apply validation to E38 and E39 (without adding labels)
        plant_hire_dv.add('E38')
        plant_hire_dv.add('E39')
        
        return sheet
    except Exception as e:
        st.error(f"Error adding dropdowns to RECOAIR sheet: {str(e)}")
        return sheet

def organize_sheets_by_area(workbook):
    """
    Reorganize the sheets in the workbook so that all related sheets for the same area 
    are grouped together in the order: CANOPY, FIRE SUPP, EBOX, RECOAIR, SDU
    """
    try:
        # Get a list of all sheets
        all_sheets = workbook.sheetnames
        
        # Create a dictionary to group sheets by area
        areas = {}
        
        # Identify all areas and their associated sheets
        for sheet_name in all_sheets:
            # Skip non-content sheets like Lists, JOB TOTAL, etc.
            if not any(prefix in sheet_name for prefix in ["CANOPY", "FIRE SUPP", "EBOX", "RECOAIR", "SDU"]):
                continue
                
            # Extract area name from sheet title
            if " - " in sheet_name:
                parts = sheet_name.split(" - ", 1)
                sheet_type = parts[0].strip()  # CANOPY, FIRE SUPP, EBOX, RECOAIR, SDU
                area_name = parts[1].strip()   # Main Kitchen (1), etc.
                
                if area_name not in areas:
                    areas[area_name] = {"CANOPY": None, "FIRE SUPP": None, "EBOX": None, "RECOAIR": None, "SDU": None}
                
                areas[area_name][sheet_type] = sheet_name
        
        # Now reorder the sheets
        # First, keep track of sheets that should not be moved (Lists, JOB TOTAL, etc.)
        static_sheets = [name for name in all_sheets if not any(prefix in name for prefix in ["CANOPY", "FIRE SUPP", "EBOX", "RECOAIR", "SDU"])]
        
        # Create the new order of sheets
        new_order = []
        
        # Put static sheets at the beginning
        for sheet_name in static_sheets:
            new_order.append(sheet_name)
        
        # Then add each area's sheets in order
        for area_name, sheet_types in areas.items():
            # Add sheets in the specified order: CANOPY, FIRE SUPP, EBOX, RECOAIR, SDU
            for sheet_type in ["CANOPY", "FIRE SUPP", "EBOX", "RECOAIR", "SDU"]:
                if sheet_types[sheet_type]:
                    new_order.append(sheet_types[sheet_type])
        
        # Apply the new order
        # Openpyxl doesn't support direct reordering, so we'll track indices
        for i, sheet_name in enumerate(new_order):
            # Get current index of this sheet
            current_index = workbook.sheetnames.index(sheet_name)
            
            # If it's not already at the desired position, move it
            if current_index != i:
                # In openpyxl, this is done by changing the _WorkbookChild.index property
                workbook._sheets.insert(i, workbook._sheets.pop(current_index))
        
        return True
    except Exception as e:
        st.error(f"Error organizing sheets: {str(e)}")
        return False

def main():
    st.set_page_config(page_title="Project Information Form", layout="wide")
    
    # Create tabs - add the new Upload Project tab
    tab1, tab2, tab3 = st.tabs(["Create New Project", "Revise Project", "Upload Project"])
    
    with tab1:
        create_general_info_form()
    
    with tab2:
        create_revision_tab()
        
    with tab3:
        create_upload_project_tab()

if __name__ == "__main__":
    main() 